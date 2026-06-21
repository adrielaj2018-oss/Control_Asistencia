# -*- coding: utf-8 -*-
"""
TAREO MÓVIL – GRUPO DE COSECHA | PRIZE PRO
Flask + SQLite/PostgreSQL + PWA. Listo para GitHub + Render.

Mejoras incluidas:
- Login con formato móvil verde/blanco como referencia.
- Pantalla principal: Soporte, Configuraciones, Sincronización y Hojas de Tareo.
- Creación de hoja: fecha, grupo, subgrupo, labor, responsable, turno y tipo.
- Detalle de hoja con tabs: Labores, Trabajadores, Rend./Avance por Labor, con iconos funcionales.
- Registro por labor-consumidor, detalle de trabajador por labor y lecturas por balde.
- Adaptado a desktop y celular con diseño responsive tipo app.
"""
import os, re, sqlite3, base64, json
from datetime import datetime, date
from functools import wraps
from io import BytesIO

from flask import Flask, request, redirect, url_for, session, flash, jsonify, send_file, render_template_string, Response
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash

try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PERSIST_DIR = os.getenv("PERSIST_DIR", "/data" if os.path.isdir("/data") else BASE_DIR)
FOTO_DIR = os.path.join(PERSIST_DIR, "fotos_marcacion")
os.makedirs(FOTO_DIR, exist_ok=True)
DB_PATH = os.path.join(PERSIST_DIR, "asistencia_tareo.db")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "cambiar-clave-en-render")
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024

# ========================= DB =========================
def is_pg(): return USE_POSTGRES and psycopg2 is not None

def get_conn():
    if is_pg():
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def qmark(sql): return sql.replace("?", "%s") if is_pg() else sql

def row_to_dict(r): return dict(r) if r else None

def rows_to_dict(rows): return [row_to_dict(r) for r in (rows or [])]

def execute(sql, params=(), fetchone=False, fetchall=False, commit=False):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(qmark(sql), params)
    data = None
    if fetchone: data = cur.fetchone()
    if fetchall: data = cur.fetchall()
    if commit: conn.commit()
    cur.close(); conn.close()
    return data

def scalar(sql, params=()):
    r = row_to_dict(execute(sql, params, fetchone=True))
    return list(r.values())[0] if r else 0

def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def today_str(): return date.today().strftime("%Y-%m-%d")


def calcular_horas_laborales(hora_inicio, hora_fin, ref_inicio=None, ref_fin=None):
    """Calcula horas netas entre inicio/fin, restando refrigerio si cruza el rango.
    Soporta turno noche cuando la hora fin es menor o igual que inicio.
    """
    def to_min(v):
        try:
            hh, mm = str(v or "00:00")[:5].split(":")
            return int(hh) * 60 + int(mm)
        except Exception:
            return 0
    ini = to_min(hora_inicio)
    fin = to_min(hora_fin)
    if fin <= ini:
        fin += 24 * 60
    total = max(0, fin - ini)
    if ref_inicio and ref_fin:
        ri = to_min(ref_inicio)
        rf = to_min(ref_fin)
        if rf <= ri:
            rf += 24 * 60
        # Si la jornada cruza medianoche y el refrigerio quedó antes del inicio, moverlo al día siguiente.
        if fin > 24 * 60 and ri < ini:
            ri += 24 * 60
            rf += 24 * 60
        cruce = max(0, min(fin, rf) - max(ini, ri))
        total -= cruce
    return round(max(0, total) / 60, 2)


def calcular_horas_nocturnas(hora_inicio, hora_fin, ref_inicio=None, ref_fin=None):
    """Horas nocturnas dentro de 22:00 a 06:00, restando refrigerio si cruza ese rango."""
    def to_min(v):
        try:
            hh, mm = str(v or "00:00")[:5].split(":")
            return int(hh) * 60 + int(mm)
        except Exception:
            return 0
    ini = to_min(hora_inicio)
    fin = to_min(hora_fin)
    if fin <= ini:
        fin += 24 * 60
    intervals = [(22*60, 30*60), (46*60, 54*60)]
    noct = 0
    for a,b in intervals:
        noct += max(0, min(fin, b) - max(ini, a))
    if ref_inicio and ref_fin:
        ri, rf = to_min(ref_inicio), to_min(ref_fin)
        if rf <= ri:
            rf += 24*60
        if fin > 24*60 and ri < ini:
            ri += 24*60; rf += 24*60
        for a,b in intervals:
            noct -= max(0, min(rf, b, fin) - max(ri, a, ini))
    return round(max(0, noct) / 60, 2)

def _minutos_hora(v):
    try:
        hh, mm = str(v or '00:00')[:5].split(':')
        return int(hh) * 60 + int(mm)
    except Exception:
        return 0

def horario_coherente(hora_inicio, hora_fin, ref_inicio, ref_fin):
    """Valida que el refrigerio esté dentro de la jornada, incluso si cruza medianoche."""
    hi = _minutos_hora(hora_inicio); hf = _minutos_hora(hora_fin)
    ri = _minutos_hora(ref_inicio); rf = _minutos_hora(ref_fin)
    if hf <= hi:
        hf += 24 * 60
    if rf <= ri:
        rf += 24 * 60
    if hf > 24 * 60 and ri < hi:
        ri += 24 * 60; rf += 24 * 60
    if not (hi <= ri < rf <= hf):
        return False, 'El refrigerio debe estar dentro de la jornada de inicio y fin.'
    if (rf - ri) > (hf - hi):
        return False, 'El refrigerio no puede ser mayor que la jornada.'
    dur_jornada = hf - hi
    dur_neta = dur_jornada - (rf - ri)
    # Bloquea horarios sin coherencia operativa: ej. 19:40 a 16:30 equivale a casi 21 horas.
    # Se mantiene permitido el turno noche normal, por ejemplo 22:00 a 06:00.
    if dur_jornada > 16 * 60 or dur_neta > 15 * 60:
        return False, 'Horario incoherente: la jornada supera el máximo permitido. Revise inicio, fin y refrigerio.'
    return True, ''

def hoja_enviada(hoja_id):
    h = row_to_dict(execute('SELECT estado FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    return bool(h and str(h.get('estado') or '').upper() == 'ENVIADA')

def _add_column_if_missing(cur, table, column, ddl):
    """Migración segura para SQLite/PostgreSQL."""
    try:
        if is_pg():
            cur.execute(qmark(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {ddl}"))
        else:
            cur.execute(f"PRAGMA table_info({table})")
            cols = [r[1] for r in cur.fetchall()]
            if column not in cols:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}")
    except Exception as e:
        print(f"No se pudo agregar columna {table}.{column}:", e)

def init_db():
    conn = get_conn(); cur = conn.cursor()
    idtype = "SERIAL PRIMARY KEY" if is_pg() else "INTEGER PRIMARY KEY AUTOINCREMENT"
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS usuarios(
        id {idtype}, usuario TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
        nombres TEXT, rol TEXT DEFAULT 'operador', estado TEXT DEFAULT 'ACTIVO', creado_en TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS trabajadores(
        id {idtype}, dni TEXT UNIQUE NOT NULL, trabajador TEXT, empresa TEXT,
        area TEXT, cargo TEXT, actividad TEXT, planilla TEXT, estado TEXT DEFAULT 'ACTIVO', fecha_carga TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS asistencia(
        id {idtype}, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        tipo TEXT NOT NULL, fecha TEXT NOT NULL, hora TEXT NOT NULL, fecha_hora TEXT NOT NULL,
        metodo TEXT, foto_path TEXT, latitud TEXT, longitud TEXT, registrado_por TEXT, observacion TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS tareos(
        id {idtype}, hoja_id INTEGER, labor_id INTEGER, dni TEXT NOT NULL, trabajador TEXT, empresa TEXT, area TEXT, cargo TEXT,
        fecha TEXT NOT NULL, labor TEXT, lote TEXT, fundo TEXT, horas REAL DEFAULT 0,
        cantidad REAL DEFAULT 0, unidad TEXT, observacion TEXT, registrado_por TEXT, creado_en TEXT,
        hora_inicio TEXT, hora_fin TEXT, ref_inicio TEXT, ref_fin TEXT, turno TEXT, tipo_tareo TEXT, horas_nocturnas REAL DEFAULT 0)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS hojas_tareo(
        id {idtype}, fecha TEXT NOT NULL, grupo TEXT, subgrupo TEXT, labor TEXT, responsable TEXT,
        turno TEXT DEFAULT 'DIA', tipo_tareo TEXT DEFAULT 'JORNAL',
        estado TEXT DEFAULT 'ABIERTA', registros INTEGER DEFAULT 0, horas_total REAL DEFAULT 0, rendimiento_total REAL DEFAULT 0,
        creado_por TEXT, creado_en TEXT, horario_fijado INTEGER DEFAULT 0, hora_inicio_default TEXT, hora_fin_default TEXT, ref_inicio_default TEXT, ref_fin_default TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS hoja_labores(
        id {idtype}, hoja_id INTEGER NOT NULL, grupo TEXT, subgrupo TEXT, labor TEXT,
        turno TEXT DEFAULT 'DIA', tipo_tareo TEXT DEFAULT 'JORNAL', responsable TEXT, creado_en TEXT, creado_por TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS actividades_maestras(
        id {idtype}, cod_actividad TEXT, desc_actividad TEXT, cod_labor TEXT, desc_labor TEXT,
        cod_consumidor TEXT, desc_consumidor TEXT, estado TEXT DEFAULT 'ACTIVO', fecha_carga TEXT)"""))
    cur.execute(qmark(f"""CREATE TABLE IF NOT EXISTS lecturas_balde(
        id {idtype}, hoja_id INTEGER, labor_id INTEGER, dni TEXT, trabajador TEXT, fecha_hora TEXT,
        a_diurno REAL DEFAULT 0, a_noct REAL DEFAULT 0, metodo TEXT, registrado_por TEXT)"""))

    # Migraciones sobre bases ya existentes en Render
    for col, ddl in [('turno', "TEXT DEFAULT 'DIA'"), ('tipo_tareo', "TEXT DEFAULT 'JORNAL'"), ('horario_fijado','INTEGER DEFAULT 0'), ('hora_inicio_default','TEXT'), ('hora_fin_default','TEXT'), ('ref_inicio_default','TEXT'), ('ref_fin_default','TEXT')]:
        _add_column_if_missing(cur, 'hojas_tareo', col, ddl)
    for col, ddl in [('labor_id','INTEGER'),('hora_inicio','TEXT'),('hora_fin','TEXT'),('ref_inicio','TEXT'),('ref_fin','TEXT'),('turno','TEXT'),('tipo_tareo','TEXT'),('horas_nocturnas','REAL DEFAULT 0')]:
        _add_column_if_missing(cur, 'tareos', col, ddl)
    for col, ddl in [('labor_id','INTEGER'),('metodo','TEXT')]:
        _add_column_if_missing(cur, 'lecturas_balde', col, ddl)

    cur.execute(qmark("SELECT id FROM usuarios WHERE usuario=?"), ("admin",))
    if not cur.fetchone():
        cur.execute(qmark("INSERT INTO usuarios(usuario,password_hash,nombres,rol,estado,creado_en) VALUES(?,?,?,?,?,?)"),
                    ("admin", generate_password_hash("admin123"), "ADMINISTRADOR", "admin", "ACTIVO", now_str()))
    conn.commit(); cur.close(); conn.close()

# ========================= UTIL =========================
def normalizar_columna(c):
    c = str(c or "").strip().upper()
    for a,b in {"Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U","Ñ":"N"}.items(): c = c.replace(a,b)
    return re.sub(r"\s+", " ", c)

def limpiar_dni(v):
    solo = re.sub(r"\D", "", str(v or ""))
    return solo[-8:] if len(solo) >= 8 else solo

def limpiar_texto(v, upper=True):
    s = "" if v is None else str(v).strip()
    return s.upper() if upper else s

def login_required(f):
    @wraps(f)
    def w(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        return f(*args, **kwargs)
    return w

def admin_required(f):
    @wraps(f)
    def w(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        if session.get("rol") != "admin":
            flash("Solo administrador puede ingresar a esta opción.", "danger")
            return redirect(url_for("home"))
        return f(*args, **kwargs)
    return w

def excel_response(headers, rows, filename, sheet="DATOS"):
    wb = Workbook(); ws = wb.active; ws.title = sheet[:31]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h.lower(), r.get(h, "")) for h in headers])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)
    out = BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========================= UI =========================
BASE_HTML = r"""
<!doctype html><html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#2f773b"><link rel="manifest" href="{{ url_for('manifest') }}">
<title>{{ title or 'Tareo Móvil' }}</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<script src="https://unpkg.com/html5-qrcode" type="text/javascript"></script>
<style>
:root{--verde:#2f773b;--verde2:#3f8748;--verde3:#276a33;--verdeClaro:#eaf5eb;--line:#e8ece8;--txt:#2d5b35;--gris:#6b7d6d;--amarillo:#ffc20e}
*{box-sizing:border-box}body{margin:0;background:#fff;font-family:Inter,Segoe UI,Arial,sans-serif;color:#21472a}.app-bg{min-height:100vh;background:linear-gradient(180deg,#fff 0%,#fff 62%,#fbfbfb 100%)}.shell{width:min(1120px,100%);margin:0 auto;padding:18px}.phone-wrap{max-width:430px;margin:0 auto}.desktop-grid{display:grid;grid-template-columns:360px 1fr;gap:28px;align-items:start}.home-desktop-list .worker-card{margin:10px 0}.home-desktop-list{max-width:640px;margin:0 auto}.header-title{text-align:center;color:#166534;font-family:Georgia,serif;font-weight:900;letter-spacing:.5px;font-size:23px;line-height:1.13;margin:4px 0 22px;text-transform:uppercase}.green-hero{background:var(--verde);border-radius:0 0 18px 18px;min-height:145px;padding:12px 16px 22px;color:white;text-align:center;position:relative;overflow:visible}.tareo-hero{min-height:124px!important;padding-bottom:42px!important}.tareo-toolbar{margin:12px 12px 8px!important;position:relative;z-index:4}.tareo-list-page .worker-card:first-of-type{margin-top:12px}.back-mini{display:inline-grid;place-items:center;width:36px;height:36px;border-radius:999px;color:var(--verde);text-decoration:none;font-size:24px}.green-top{display:flex;justify-content:space-between;align-items:center;font-size:11px;font-weight:800}.avatar{width:78px;height:78px;border-radius:999px;background:white;color:var(--verde);display:grid;place-items:center;margin:10px auto 2px;font-size:43px;box-shadow:0 8px 20px rgba(0,0,0,.13)}.login-name{font-size:11px;font-weight:800}.white-input{height:36px;background:white;border-radius:10px;box-shadow:0 5px 13px rgba(0,0,0,.18);border:0}.floating-card{background:white;border-radius:10px;box-shadow:0 8px 18px rgba(0,0,0,.15);padding:12px}.tile{width:74px;height:70px;border-radius:8px;background:white;box-shadow:0 7px 17px rgba(0,0,0,.14);display:flex;flex-direction:column;align-items:center;justify-content:center;color:var(--verde);font-weight:900;font-size:10px;text-align:center}.tile i{font-size:26px;margin-bottom:5px}.bottom-sync{position:fixed;left:10px;bottom:8px;color:#317a3e;font-size:10px;font-weight:700}.bottom-out{position:fixed;right:14px;bottom:8px;color:#c84c4c;font-size:20px}.tab-main{display:flex;gap:0;background:#fafafa;padding-left:0;border-top:4px solid #d7d7d7}.tab-main a{flex:1;text-align:center;text-decoration:none;color:#508557;font-weight:900;font-size:13px;padding:14px 8px;border-radius:7px 7px 0 0;background:#fff}.tab-main a.active{background:var(--verde);color:white;box-shadow:0 3px 7px rgba(0,0,0,.18)}.subtabs{display:flex;background:#fff}.subtabs a{flex:1;text-align:center;padding:13px 5px;text-decoration:none;color:#4b8a54;font-weight:900;font-size:12px}.subtabs a.active{background:var(--verde);color:white}.panel-green{background:var(--verde);color:white;text-align:center;padding:21px 12px 42px}.panel-green i{font-size:38px}.panel-green h4{font-size:11px;font-weight:900;margin:5px 0 0}.toolstrip{background:white;margin:-25px 9px 5px;border-radius:9px;min-height:49px;box-shadow:0 5px 13px rgba(0,0,0,.22);display:flex;align-items:center;gap:20px;padding:7px 14px;color:var(--verde);font-size:24px}.toolstrip button,.toolstrip a{border:0;background:transparent;color:var(--verde);font-size:24px;text-decoration:none}.info-bar{margin:0 9px;background:var(--verde);color:white;border-radius:2px;display:grid;grid-template-columns:1fr 1fr 1fr 1fr 22px;align-items:center;font-size:10px;font-weight:900;height:23px}.info-bar div{text-align:center;border-right:1px solid rgba(255,255,255,.28)}.worker-card{background:white;margin:10px 12px;border-radius:10px;border:1px solid #e3e8e3;box-shadow:0 3px 12px rgba(0,0,0,.20);padding:11px 13px;color:#397443;position:relative}.worker-title{display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:9px;font-weight:900;text-transform:uppercase}.worker-title b{font-size:10px;color:var(--verde)}.worker-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:7px;margin-top:7px}.worker-grid label{font-size:8px;font-weight:900;color:#6c8a6f;margin-bottom:1px}.mini-input{height:25px;border:1px solid #9ebaa0;border-radius:3px;font-size:10px;padding:3px 5px;width:100%;font-weight:800;color:#315f39}.mini-badge{border-radius:3px;color:white;font-size:8px;font-weight:900;text-align:center;padding:4px 3px;text-transform:uppercase}.bg-y{background:var(--amarillo)!important;color:white}.bg-g{background:#42b852!important}.person-dot{width:42px;height:42px;border-radius:999px;background:#378145;color:white;display:grid;place-items:center;font-size:26px;float:left;margin-right:9px}.small-label{font-size:8px;color:#79937b;font-weight:900}.small-value{font-size:9px;color:#466a49;font-weight:900}.leaf{width:120px;height:120px;border-radius:70% 30% 70% 30%;background:linear-gradient(135deg,#ffd8bd,#eef4c7,#cbd9b6);opacity:.72;margin:20px auto 0;transform:rotate(-20deg)}.card-pro{background:white;border:1px solid var(--line);border-radius:18px;box-shadow:0 8px 20px rgba(0,0,0,.10)}.btn-green{background:var(--verde);border-color:var(--verde);color:white;font-weight:900;border-radius:9px}.btn-green:hover{background:var(--verde3);color:white}.form-control,.form-select{border-radius:9px;border:1px solid #dfe7df;font-weight:700;font-size:13px}.form-label{font-size:12px;font-weight:900;color:#3e7545}.page-card{border-radius:13px;overflow:hidden;border:1px solid #e5e7e5;background:white}.list-table th{font-size:11px;color:#497550}.list-table td{font-size:12px;vertical-align:middle}.status-pill{display:inline-block;background:#39b54a;color:white;border-radius:4px;padding:4px 8px;font-size:9px;font-weight:900}.top-actions{display:flex;gap:12px;flex-wrap:wrap;justify-content:center;margin-top:14px;position:relative;z-index:5}.top-actions .tile{width:82px;height:76px}.login-page .shell{padding:0}.login-form{margin:-7px auto 0;width:92%;max-width:360px}.login-form .floating-card{padding:13px 14px 18px}.alert{border-radius:12px;font-size:13px}.desk-panel{display:block}.mobile-only{display:none}.clock-box{width:116px;height:116px;border:5px solid var(--verde);border-radius:999px;margin:8px auto;display:grid;place-items:center;color:var(--verde);font-weight:900;background:#fff;box-shadow:0 4px 14px rgba(47,119,59,.18)}.clock-box i{font-size:38px}.scan-box{border:2px dashed #8dbf93;border-radius:12px;padding:10px;background:#f8fff9}.toolstrip .hint{font-size:9px;font-weight:900;color:#2f773b;margin-left:-18px;margin-right:0}.splash-card{height:94vh;max-height:760px;background:#23773f;border-radius:10px;box-shadow:0 4px 12px rgba(0,0,0,.25);display:flex;flex-direction:column;align-items:center;justify-content:center;color:white;position:relative}.splash-logo{width:145px;height:145px;border-radius:999px;background:#fff;border:6px solid #92bd33;display:grid;place-items:center;color:#23773f;font-size:66px;box-shadow:0 3px 10px rgba(0,0,0,.22)}.splash-title{font-weight:900;margin-top:18px;letter-spacing:.5px}.splash-foot{position:absolute;bottom:26px;text-align:center;font-size:11px;color:#d9f2df;font-weight:700}.role-toggle{display:grid;grid-template-columns:1fr 1fr;gap:8px}.role-toggle label{border:1px solid #dce7dc;border-radius:9px;padding:9px;text-align:center;font-size:12px;font-weight:900;color:#2f773b}.role-toggle input{display:none}.role-toggle input:checked+span{background:#2f773b;color:white;border-radius:7px;padding:7px 9px;display:block}.bottom-nav{position:sticky;bottom:0;background:white;border-top:1px solid #e8ece8;display:flex;justify-content:space-around;padding:7px 0;color:#477b4d;font-size:10px;font-weight:800}.bottom-nav a{text-decoration:none;color:#477b4d;text-align:center}.bottom-nav i{display:block;font-size:17px}.copy-list{max-height:260px;overflow:auto;border:1px solid #e5ede5;border-radius:9px;padding:8px;background:#fbfffb}.clock-face{width:180px;height:180px;border-radius:999px;background:#e7e0ef;margin:8px auto;position:relative;display:grid;place-items:center;color:#5d44aa;touch-action:none;cursor:pointer;user-select:none}.clock-hand{width:65px;height:4px;background:#6b4eb8;position:absolute;transform-origin:left center;transform:rotate(-35deg);left:90px;top:90px;pointer-events:none}.clock-hand:after{content:'';position:absolute;right:-13px;top:-13px;width:30px;height:30px;border-radius:999px;background:#6b4eb8;box-shadow:0 2px 8px rgba(0,0,0,.18)}.clock-dot{width:12px;height:12px;border-radius:999px;background:#6b4eb8;position:absolute;left:84px;top:84px}.clock-num{position:absolute;font-size:12px;color:#37303c}.clock-bubble{position:absolute;right:20px;top:50px;background:#6b4eb8;color:white;border-radius:999px;padding:9px;font-weight:900}.field-required{box-shadow:inset 4px 0 0 var(--verde)}.big-plus{font-size:34px!important;line-height:1}.big-plus .bi-plus{font-size:22px!important;margin-left:-12px;font-weight:900}.labor-card-compact{padding:13px 16px}.labor-card-compact .worker-title{font-size:8px}.labor-card-compact .worker-title b{font-size:9px;line-height:1.15}.labor-card-compact .labor-main{font-size:15px!important;line-height:1.1;color:#146c35}.labor-card-compact .resp-main{font-size:13px!important;line-height:1.1;color:#146c35}.worker-queue{border:1px dashed #8cc79b;border-radius:12px;background:#f7fff8;padding:9px;margin-top:10px;max-height:155px;overflow:auto}.queue-item{display:flex;justify-content:space-between;gap:8px;align-items:center;border-bottom:1px solid #e2f3e5;padding:6px 0;font-size:12px}.queue-item:last-child{border-bottom:0}.scan-ok{background:#d1fae5;border:1px solid #86efac;color:#166534;border-radius:10px;padding:8px;font-size:12px;font-weight:800}.scan-bad{background:#fee2e2;border:1px solid #fecaca;color:#991b1b;border-radius:10px;padding:8px;font-size:12px;font-weight:800}.time-click{cursor:pointer;background:#fbfffb}.time-click:focus{outline:2px solid #2f773b}.report-wrap{max-width:540px;margin:0 auto}.config-header{display:flex;align-items:center;gap:8px;justify-content:center;position:relative}.config-header .back-mini{position:absolute;left:0}.btn-plus-fab{display:inline-flex!important;align-items:center;gap:2px}.btn-plus-fab i:first-child{font-size:30px!important}.btn-plus-fab i:last-child{font-size:20px!important;margin-left:-12px;margin-top:12px}.field-help{font-size:10px;color:#5f7d65;font-weight:800} .swipe-wrap{position:relative;margin:10px 12px;overflow:hidden;border-radius:12px;border:1px solid #e3e8e3;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,.06)}.swipe-actions{position:absolute;right:0;top:0;bottom:0;display:flex;align-items:stretch;transform:translateX(100%);transition:.22s ease;z-index:1}.swipe-wrap.show-actions .swipe-actions{transform:translateX(0)}.swipe-actions a{display:flex;align-items:center;justify-content:center;min-width:74px;color:white;text-decoration:none;font-size:11px;font-weight:900}.act-edit{background:#2563eb}.act-send{background:#16a34a}.act-del{background:#dc2626}.swipe-wrap .worker-card{margin:0;transition:.22s ease;position:relative;z-index:2}.swipe-wrap.show-actions .worker-card{transform:translateX(-222px)}.locked-input{background:#f8fff9!important;cursor:pointer}.edit-hint{font-size:9px;color:#2f773b;font-weight:900}.clock-field-pills{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:4px 0 10px}.clock-field-pills button{border:1px solid #cfe6d4;background:#fff;border-radius:9px;padding:7px;font-size:11px;font-weight:900;color:#2f773b}.clock-field-pills button.active{background:#2f773b;color:white}.clock-mode{display:flex;justify-content:center;gap:8px;margin-top:4px}.clock-mode button{border:0;border-radius:999px;padding:5px 12px;font-weight:900;background:#e7e0ef;color:#5d44aa}.clock-mode button.active{background:#6b4eb8;color:white}.modal-suggest{max-height:125px;overflow:auto;border:1px solid #e5ede5;border-radius:8px;margin-top:-5px;margin-bottom:8px;background:#fbfffb;display:none}.modal-suggest div{padding:8px 10px;border-bottom:1px solid #edf5ee;font-size:12px;font-weight:800;cursor:pointer}.modal-suggest div:hover{background:#eaf5eb}.queue-title{font-size:11px;font-weight:900;color:#166534;margin-top:8px}.scan-ok.flash{animation:flashOk .5s ease}@keyframes flashOk{0%{transform:scale(.98)}50%{transform:scale(1.02)}100%{transform:scale(1)}} 

/* ===== MEJORAS 246: mayúsculas, reloj táctil y layout horas ===== */
input[type="text"], input:not([type]), textarea { text-transform:uppercase; }
.time-worker-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}
.time-worker-grid label{font-size:8px;font-weight:900;color:#55745a;margin-bottom:2px;display:block}
.time-box{background:#fbfffb;border:1px solid #9ebaa0;border-radius:5px;padding:5px 6px;font-size:11px;font-weight:900;color:#245a31;min-height:29px}
.time-metrics{display:grid;grid-template-columns:1fr;gap:7px;margin-top:8px;background:#eef8ef;border:1px solid #b9d7bd;border-radius:8px;padding:8px}
.time-metrics label{font-size:8px;font-weight:900;color:#176a35;margin-bottom:2px;display:block}
.metric-box{background:#dff3e2;border:1px solid #81ba8a;border-radius:6px;padding:7px 8px;font-size:13px;font-weight:900;color:#0f6b2c;min-height:32px}
.ref-worker-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}
.ref-worker-grid label{font-size:8px;font-weight:900;color:#55745a;margin-bottom:2px;display:block}
.clock-face.pro-clock{width:210px;height:210px;background:#f4f8f4;border:7px solid #2f773b;color:#2f773b;box-shadow:0 6px 18px rgba(47,119,59,.2)}
.clock-face.pro-clock .clock-hand{left:105px;top:105px;background:#2f773b;width:76px;height:5px}
.clock-face.pro-clock .clock-hand:after{background:#2f773b}
.clock-face.pro-clock .clock-dot{left:99px;top:99px;background:#2f773b}
.clock-face.pro-clock .clock-bubble{background:#2f773b}
.clock-24-hint{text-align:center;font-size:11px;font-weight:900;color:#2f773b;margin-top:-4px;margin-bottom:6px}


/* === AJUSTE FINAL OMAR: modal compacto, tarjetas como referencia y controles táctiles/PC === */
.modal-dialog{max-width:520px!important;margin:.7rem auto!important}.modal-content{border-radius:12px!important}.modal-header{padding:10px 16px!important}.modal-body{padding:14px 16px!important}.modal-footer{padding:10px 16px!important}.modal .form-label{font-size:11px!important;margin-bottom:4px!important}.modal .form-control,.modal .form-select{height:41px!important;font-size:13px!important;border-radius:9px!important}.modal .btn-green{height:47px!important;border-radius:9px!important}.modal .alert{padding:10px 12px!important;margin-bottom:10px!important}
#modalLabor .modal-dialog{max-width:500px!important}#modalLabor .modal-body{padding-top:12px!important}#modalLabor .form-control,#modalLabor .form-select{height:40px!important}#modalLabor .modal-suggest{position:relative;z-index:2000;max-height:160px;overflow:auto;background:#fff;border:1px solid #b8d9bf;border-radius:8px;margin:2px 0 8px;box-shadow:0 8px 20px rgba(0,0,0,.10);display:none}#modalLabor .modal-suggest div{padding:9px 10px;font-size:12px;font-weight:800;color:#245c31;border-bottom:1px solid #edf5ee;cursor:pointer}#modalLabor .modal-suggest div:hover{background:#eaf5eb}.master-status{font-size:10px;font-weight:900;color:#2f773b;margin:-2px 0 6px}.master-status.bad{color:#b42318}
#modalHora .modal-dialog{max-width:490px!important}.touch-clock-panel{padding:10px!important;border-radius:12px!important}.time-display{height:42px!important;font-size:28px!important;line-height:39px!important}.time-slider{width:100%!important;height:34px!important;display:block!important;cursor:pointer!important;touch-action:pan-x!important;accent-color:#2f773b!important}.time-slider::-webkit-slider-thumb{width:28px!important;height:28px!important;cursor:pointer!important}.touch-clock-picks{grid-template-columns:1fr 1fr!important;gap:7px!important}.touch-clock-picks button{height:35px!important;border-radius:8px!important;font-size:10px!important}.locked-input{pointer-events:auto!important;cursor:pointer!important;background:#f8fff9!important}.trabajador-card-ref{padding:20px 25px!important;border-radius:13px!important;max-width:790px;margin:10px auto!important}.trabajador-card-ref .worker-title{grid-template-columns:1fr 1fr!important;font-size:12px!important;gap:35px!important}.trabajador-card-ref .worker-title b{font-size:18px!important;line-height:1.15!important}.trabajador-grid-ref{display:grid;grid-template-columns:1fr 1fr 1fr;gap:26px 20px;margin-top:34px}.trabajador-grid-ref label{font-size:13px;font-weight:900;color:#6b7d6d;margin-bottom:8px;display:block}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:48px!important;display:flex;align-items:center;border:1px solid #9ebaa0;border-radius:4px;background:#fbfffb;color:#006b2e;font-weight:900;font-size:22px;padding:8px 12px}.trabajador-grid-ref .mini-badge{height:48px;display:flex;align-items:center;justify-content:center;font-size:15px;border-radius:4px;max-width:236px!important}.editable-tareo{cursor:pointer}.editable-tareo:hover{outline:2px solid #2f773b}
@media(max-width:860px){.modal-dialog{max-width:94%!important;margin:.65rem auto!important}.trabajador-card-ref{padding:12px 14px!important;max-width:505px!important}.trabajador-card-ref .worker-title{font-size:9px!important;gap:10px!important}.trabajador-card-ref .worker-title b{font-size:13px!important}.trabajador-grid-ref{gap:12px 8px;margin-top:18px}.trabajador-grid-ref label{font-size:9px;margin-bottom:5px}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:35px!important;font-size:15px!important;padding:5px 8px}.trabajador-grid-ref .mini-badge{height:36px;font-size:11px;max-width:185px!important}}

@media(max-width:860px){.shell{padding:0}.desktop-grid{display:block}.desk-panel{display:none}.mobile-only{display:block}.header-title{font-size:17px;margin:16px 7px 20px}.page-card{border-radius:0;border-left:0;border-right:0}.phone-wrap{max-width:100%}.green-hero{border-radius:0}.worker-card{margin-left:9px;margin-right:9px}.toolstrip{gap:15px}.info-bar{font-size:8.5px}.bottom-sync,.bottom-out{position:fixed}.desktop-pad{padding:0 0 28px}.tab-main a,.subtabs a{font-size:11px}.worker-grid{gap:5px}.floating-card{border-radius:9px}.top-actions .tile{width:72px;height:70px}}

/* ===== AJUSTE FINAL: horario táctil compacto y edición de tareo ===== */
.clock-face.pro-clock{display:none!important}
.touch-clock-panel{background:#f8fff9;border:1px solid #cfe6d4;border-radius:10px;padding:8px;margin:6px 0 10px}
.touch-clock-picks{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px}
.touch-clock-picks button{border:1px solid #b9d7bd;background:#fff;border-radius:8px;padding:7px 4px;font-size:10px;font-weight:900;color:#2f773b}
.touch-clock-picks button.active{background:#2f773b;color:white}
.touch-clock-actions{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:6px}
.touch-clock-actions button{border:0;background:#2f773b;color:white;border-radius:8px;padding:8px 2px;font-size:11px;font-weight:900}
.touch-clock-value{text-align:center;font-size:24px;font-weight:900;color:#166534;background:#fff;border:1px solid #cfe6d4;border-radius:10px;padding:6px;margin-bottom:8px}
.modal-dialog{max-width:390px}.modal-content{border-radius:13px}.modal-header{padding:10px 14px}.modal-body{padding:12px 14px}.modal-footer{padding:10px 14px}.modal-title{font-size:18px}.modal .form-label{font-size:11px;margin-bottom:3px}.modal .form-control,.modal .form-select{height:37px;font-size:12px;padding:6px 10px}.modal .alert{font-size:12px;padding:9px 11px;margin-bottom:10px}
.time-slider{width:100%;accent-color:#2f773b;touch-action:pan-y;margin:6px 0 3px}.time-display{height:33px;border:1px solid #cfe6d4;border-radius:9px;background:#fff;display:grid;place-items:center;font-weight:900;color:#166534;font-size:20px;margin-bottom:7px}.touch-clock-panel{padding:7px!important;margin:4px 0 8px!important}.touch-clock-picks{gap:5px!important;margin-bottom:6px!important}.touch-clock-picks button{padding:6px 3px!important;font-size:9px!important}.touch-clock-actions{display:none!important}
.time-worker-grid,.ref-worker-grid{gap:6px}.time-metrics{padding:7px;gap:6px}.worker-card{padding:9px 11px}.metric-box{padding:5px 7px;min-height:28px}.time-box{min-height:27px;padding:4px 6px}
.worker-card.editable-tareo{cursor:pointer}.worker-card.editable-tareo:after{content:'Tocar para editar';position:absolute;right:10px;bottom:8px;font-size:8px;font-weight:900;color:#2f773b}


/* === PARCHE REAL 247: tarjeta trabajador en 3 columnas compacto === */
.trabajador-card-ref{max-width:790px!important;padding:18px 24px!important}
.trabajador-grid-ref{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:22px 18px!important;margin-top:28px!important}
.trabajador-grid-ref label{font-size:12px!important;margin-bottom:7px!important}
.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:46px!important;font-size:20px!important;background:#fbfffb!important;border:1px solid #9ebaa0!important;border-radius:4px!important;display:flex!important;align-items:center!important;color:#006b2e!important;font-weight:900!important}
@media(max-width:860px){.trabajador-card-ref{max-width:505px!important;padding:10px 13px!important}.trabajador-grid-ref{grid-template-columns:repeat(3,1fr)!important;gap:10px 7px!important;margin-top:16px!important}.trabajador-grid-ref label{font-size:8.5px!important;margin-bottom:4px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:32px!important;font-size:13px!important;padding:4px 7px!important}.trabajador-card-ref .worker-title b{font-size:12px!important}}

/* ===== PATCH 248 OMAR: sin doble desplegable, labor obligatoria, tarjeta compacta y slider tipo manija ===== */
#modalLabor datalist{display:none!important}
#modalLabor .modal-dialog{max-width:390px!important}
#modalLabor .modal-body{padding:10px 14px!important}
#modalLabor .form-control,#modalLabor .form-select{height:36px!important;font-size:12px!important}
#modalLabor .modal-suggest{position:absolute!important;left:20px!important;right:20px!important;z-index:4000!important;max-height:142px!important;overflow:auto!important;background:#fff!important;box-shadow:0 10px 24px rgba(0,0,0,.18)!important}
#modalHora .modal-dialog{max-width:430px!important}.touch-clock-panel{padding:7px!important}.time-display{height:36px!important;font-size:22px!important}.time-slider{height:38px!important;cursor:pointer!important;touch-action:none!important;appearance:none!important;-webkit-appearance:none!important;background:transparent!important}.time-slider::-webkit-slider-runnable-track{height:8px;background:#dfe7df;border:1px solid #9ebaa0;border-radius:999px}.time-slider::-webkit-slider-thumb{-webkit-appearance:none!important;width:38px!important;height:38px!important;margin-top:-16px!important;border-radius:999px!important;background:#2f773b!important;border:4px solid #fff!important;box-shadow:0 3px 9px rgba(0,0,0,.28)!important;cursor:grab!important}.time-slider::-webkit-slider-thumb:active{cursor:grabbing!important}.time-slider::-moz-range-track{height:8px;background:#dfe7df;border:1px solid #9ebaa0;border-radius:999px}.time-slider::-moz-range-thumb{width:34px!important;height:34px!important;border-radius:999px!important;background:#2f773b!important;border:4px solid #fff!important;box-shadow:0 3px 9px rgba(0,0,0,.28)!important;cursor:grab!important}.touch-clock-panel:after{content:'◷';display:block;text-align:center;color:#2f773b;font-size:18px;font-weight:900;margin-top:-4px}.trabajador-card-ref{max-width:100%!important;margin:7px 8px!important;padding:10px 14px!important;border-radius:12px!important}.trabajador-card-ref .worker-title{gap:8px!important;font-size:8px!important}.trabajador-card-ref .worker-title b{font-size:13px!important;line-height:1.12!important}.trabajador-grid-ref{grid-template-columns:repeat(3,1fr)!important;gap:10px 8px!important;margin-top:14px!important}.trabajador-grid-ref label{font-size:8.5px!important;margin-bottom:4px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:31px!important;font-size:14px!important;padding:4px 7px!important;border-radius:4px!important}.trabajador-grid-ref .mini-badge{height:35px!important;font-size:11px!important;max-width:100%!important}.worker-card.editable-tareo:after{right:10px!important;bottom:5px!important;font-size:8px!important}.phone-wrap{max-width:560px!important}@media(max-width:860px){.phone-wrap{max-width:100%!important}.trabajador-card-ref{margin:7px 5px!important;padding:9px 12px!important}.trabajador-card-ref .worker-title b{font-size:12px!important}.trabajador-grid-ref{gap:9px 7px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:30px!important;font-size:13px!important}}


/* PATCH URGENTE: coherencia horario, avance auto, registros compactos */
.trabajador-card-ref{padding:7px 10px!important;margin:5px 4px!important;border-width:2px!important}.trabajador-card-ref .worker-title b{font-size:12px!important}.trabajador-grid-ref{gap:7px 6px!important;margin-top:9px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:28px!important;justify-content:center!important;text-align:center!important;font-size:12px!important}.trabajador-grid-ref .metric-box{font-size:14px!important;background:#daf5dc!important}.trabajador-grid-ref label{font-size:7.5px!important}.avance-card-ref{padding:8px 10px!important;margin:6px 10px!important}.avance-card-ref .worker-title b{font-size:12px!important}.avance-grid .mini-badge{height:26px!important;display:flex;align-items:center;justify-content:center}.labor-card-compact{padding:10px 13px!important}.labor-card-compact .labor-main{font-size:16px!important}.labor-card-compact .resp-main{font-size:14px!important}.page-card{max-width:620px!important;margin:0 auto!important}.phone-wrap{max-width:650px!important}#modalHora input.locked-input{background:#fff!important;cursor:text!important}.scan-ok{background:#e7f7ea!important;border-color:#8ad092!important}.scan-bad{background:#fee2e2!important;border-color:#fca5a5!important;color:#991b1b!important}
@media(max-width:720px){.phone-wrap{max-width:100%!important}.page-card{max-width:100%!important}.trabajador-card-ref{padding:6px 8px!important}.trabajador-grid-ref{gap:6px 5px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:26px!important;font-size:11px!important}.trabajador-grid-ref .metric-box{font-size:13px!important}.avance-card-ref{margin:5px 8px!important}}


/* === PATCH 249 FINAL: ancho tipo celular + tarjetas compactas + flecha acciones === */
html, body{max-width:100%;overflow-x:hidden!important;background:#fff!important;}
.shell{width:100%!important;max-width:430px!important;margin:0 auto!important;padding:6px 8px!important;}
.phone-wrap{width:100%!important;max-width:390px!important;margin:0 auto!important;}
.page-card{width:100%!important;max-width:390px!important;margin:0 auto!important;border-radius:12px!important;overflow:hidden!important;}
.header-title{font-size:16px!important;line-height:1.08!important;margin:10px 4px 12px!important;letter-spacing:.2px!important;}
.green-hero,.panel-green{padding:13px 10px 28px!important;min-height:104px!important;border-radius:0!important;}
.panel-green i{font-size:28px!important}.panel-green h4{font-size:10px!important;line-height:1.15!important;margin-top:4px!important;}
.tab-main a{font-size:11px!important;padding:10px 4px!important}.subtabs a{font-size:10px!important;padding:8px 4px!important;line-height:1.15!important;}
.toolstrip{margin:-22px 8px 5px!important;min-height:42px!important;padding:6px 11px!important;gap:16px!important;border-radius:8px!important;}
.toolstrip button,.toolstrip a{font-size:21px!important}.info-bar{margin:0 8px!important;height:22px!important;font-size:9px!important;grid-template-columns:1fr 1fr 1fr 1fr 18px!important;}
.leaf{width:80px!important;height:80px!important;margin:14px auto 0!important;opacity:.55!important;}
.worker-card{margin:7px 7px!important;padding:8px 9px!important;border-radius:9px!important;box-shadow:0 4px 11px rgba(0,0,0,.13)!important;}
.labor-card-compact{padding:8px 10px!important}.labor-card-compact .worker-title{font-size:7.5px!important}.labor-card-compact .worker-title b{font-size:9px!important}.labor-card-compact .labor-main{font-size:14px!important}.labor-card-compact .resp-main{font-size:12px!important}.labor-card-compact .mini-badge{height:22px!important;font-size:8px!important;padding:4px!important;}
.trabajador-card-ref{width:auto!important;max-width:100%!important;margin:6px 5px!important;padding:7px 9px 18px!important;border-radius:10px!important;border-width:2px!important;}
.trabajador-card-ref .worker-title{grid-template-columns:1.25fr .85fr!important;gap:5px!important;font-size:7px!important;line-height:1.1!important;}
.trabajador-card-ref .worker-title b{font-size:10.5px!important;line-height:1.12!important;}
.trabajador-grid-ref{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:6px 6px!important;margin-top:9px!important;}
.trabajador-grid-ref label{font-size:7px!important;margin-bottom:2px!important;line-height:1!important;}
.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:27px!important;font-size:11px!important;padding:3px 5px!important;justify-content:center!important;text-align:center!important;border-radius:4px!important;}
.trabajador-grid-ref .metric-box{font-size:13px!important;background:#dff6df!important;border:1px solid #42b852!important;color:#006b2e!important;}
.trabajador-grid-ref div:has(label+ .metric-box):last-child .metric-box,.trabajador-grid-ref div:nth-child(6) .metric-box{background:#dbeafe!important;border-color:#60a5fa!important;color:#0057a8!important;}
.trabajador-grid-ref .mini-badge{height:30px!important;font-size:10px!important;max-width:100%!important;}
.avance-card-ref{margin:6px 7px!important;padding:8px 9px!important;border-radius:9px!important;min-height:82px!important;}
.avance-card-ref .person-dot{width:34px!important;height:34px!important;font-size:20px!important;margin-right:8px!important}.avance-card-ref .worker-title{font-size:7.5px!important}.avance-card-ref .worker-title b{font-size:10px!important}.avance-grid{grid-template-columns:1fr 1fr 32px!important;align-items:end!important;gap:6px!important}.avance-grid .mini-badge{height:24px!important;font-size:9px!important;display:flex!important;align-items:center!important;justify-content:center!important;}
.card-action-chevron{position:absolute;right:8px;top:50%;transform:translateY(-50%);width:28px;height:38px;display:grid;place-items:center;color:#08713b;font-size:28px;font-weight:900;cursor:pointer;z-index:10;border-radius:8px;background:rgba(255,255,255,.75)}
.card-action-chevron:hover{background:#eaf7ed}.card-action-chevron::before{content:'‹';}
.card-menu{display:none;position:absolute;right:36px;top:50%;transform:translateY(-50%);background:#fff;border:1px solid #b9d7bd;border-radius:9px;box-shadow:0 8px 20px rgba(0,0,0,.18);z-index:20;overflow:hidden;min-width:112px;}
.card-menu.show{display:block}.card-menu button,.card-menu a{display:block;width:100%;padding:9px 10px;border:0;background:#fff;text-decoration:none;text-align:left;color:#166534;font-size:11px;font-weight:900}.card-menu .danger{color:#b42318}.card-menu button:hover,.card-menu a:hover{background:#eef8ef;}
.worker-card.editable-tareo:after{display:none!important;content:''!important;}
/* formulario crear hoja y modales encajados */
body .modal-dialog{max-width:365px!important;margin:.55rem auto!important;}.modal-content{border-radius:12px!important}.modal-header{padding:8px 12px!important}.modal-body{padding:10px 12px!important}.modal-footer{padding:8px 12px!important}.modal-title{font-size:17px!important}.modal .alert{font-size:11px!important;padding:8px 10px!important}.modal .form-label{font-size:10px!important;margin-bottom:3px!important}.modal .form-control,.modal .form-select{height:34px!important;font-size:12px!important;border-radius:8px!important;padding:5px 9px!important}.modal .btn-green{height:40px!important;font-size:13px!important;border-radius:8px!important;}
#createHojaCompact .panel-green{padding:14px 10px 24px!important;min-height:90px!important}#createHojaCompact .floating-card{margin:-18px 7px 10px!important;padding:10px!important}#createHojaCompact .form-control,#createHojaCompact .form-select{height:36px!important;font-size:12px!important}#createHojaCompact .btn{height:38px!important;font-size:13px!important;padding:6px 10px!important}.desktop-pad{padding:0 0 18px!important;}
/* horario: recuadros de hora deslizables y digitables */
#modalHora .touch-clock-panel{padding:7px!important;margin:4px 0 8px!important}#modalHora .clock-face{width:165px!important;height:165px!important}#modalHora .time-display{height:34px!important;font-size:21px!important}.time-drag-help{font-size:9px;color:#166534;font-weight:900;margin-top:3px}.time-draggable{cursor:ew-resize!important;touch-action:none!important;background:#fff!important;}
@media(max-width:420px){.shell{padding:4px 6px!important}.phone-wrap,.page-card{max-width:360px!important}.tab-main a{font-size:10px!important}.subtabs a{font-size:9px!important}.toolstrip{gap:13px!important}.trabajador-card-ref .worker-title b{font-size:9.5px!important}.trabajador-grid-ref .time-box,.trabajador-grid-ref .metric-box{height:25px!important;font-size:10.5px!important}.trabajador-grid-ref .metric-box{font-size:12px!important}.modal-dialog{max-width:94%!important}}


/* PATCH 252: botón visible para apagar cámara */
.scanner-close-x{position:absolute!important;right:8px!important;top:8px!important;z-index:99999!important;width:36px!important;height:36px!important;border:0!important;border-radius:999px!important;background:#dc2626!important;color:#fff!important;font-size:24px!important;font-weight:900!important;line-height:34px!important;text-align:center!important;box-shadow:0 3px 10px rgba(0,0,0,.28)!important;}
.scanner-close-x:hover{background:#b91c1c!important;}
[id^="reader"]{position:relative!important;min-height:0;}


/* PATCH 253: selector táctil tipo iOS para horas HH:MM */
.ios-time-picker{position:relative;display:grid;grid-template-columns:1fr 1fr;gap:8px;height:154px;margin:8px 0 10px;padding:42px 10px;background:#f7fff8;border:1px solid #cfe6d4;border-radius:14px;overflow:hidden;box-shadow:inset 0 1px 0 rgba(255,255,255,.8)}
.ios-time-picker:before{content:'';position:absolute;left:10px;right:10px;top:60px;height:34px;background:rgba(47,119,59,.10);border:1px solid #b7d8bd;border-radius:9px;pointer-events:none;z-index:1}
.ios-time-picker:after{content:':';position:absolute;left:50%;top:58px;transform:translateX(-50%);font-size:27px;font-weight:900;color:#166534;z-index:2;pointer-events:none}
.ios-wheel{height:70px;overflow-y:auto;scroll-snap-type:y mandatory;-webkit-overflow-scrolling:touch;overscroll-behavior:contain;text-align:center;z-index:3;scrollbar-width:none;background:transparent;border:0}
.ios-wheel::-webkit-scrollbar{display:none}.ios-wheel .ios-pad{height:18px}.ios-wheel button{display:block;width:100%;height:34px;border:0;background:transparent;scroll-snap-align:center;font-size:20px;font-weight:900;color:#5f6b62;line-height:34px}.ios-wheel button.active{color:#006b2e;font-size:23px}.ios-time-label{text-align:center;font-size:10px;font-weight:900;color:#166534;margin-top:-4px;margin-bottom:3px}.ios-time-hint{font-size:9px;color:#517a58;text-align:center;font-weight:800;margin-top:-5px;margin-bottom:4px}

</style></head><body class="{{ 'login-page' if not session.get('usuario') else '' }}"><div class="app-bg"><main class="shell">
{% with messages=get_flashed_messages(with_categories=true) %}{% if messages %}<div class="phone-wrap mt-2">{% for cat,msg in messages %}<div class="alert alert-{{cat}} shadow-sm">{{msg}}</div>{% endfor %}</div>{% endif %}{% endwith %}
{{ body|safe }}</main></div>
<audio id="sndOk"><source src="data:audio/wav;base64,UklGRjQAAABXQVZFZm10IBAAAAABAAEAESsAACJWAAACABAAZGF0YRAAAAAAAP//AAD//wAA//8AAP//AAD//wAA//8=" type="audio/wav"></audio>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script><script>function beep(){try{let a=document.getElementById('sndOk');a.currentTime=0;a.play().catch(()=>{});}catch(e){}}if('serviceWorker' in navigator){navigator.serviceWorker.register('/sw.js').catch(()=>{});}function bindSwipeCards(){document.querySelectorAll('.swipe-wrap').forEach(w=>{let sx=0,dx=0;w.addEventListener('touchstart',e=>{sx=e.touches[0].clientX;});w.addEventListener('touchmove',e=>{dx=e.touches[0].clientX-sx;});w.addEventListener('touchend',()=>{if(dx<-35)w.classList.add('show-actions'); if(dx>35)w.classList.remove('show-actions'); dx=0;});w.addEventListener('contextmenu',e=>{e.preventDefault();w.classList.toggle('show-actions');});});}document.addEventListener('DOMContentLoaded',bindSwipeCards);
</script>

<script>
/* PARCHE FINAL DE LECTURA AUTOMÁTICA DNI/QR/BARRAS - no depende del script del modal */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  let timer=null, busy=false, last='';
  window.__dniQueueFinal = window.__dniQueueFinal || new Map();
  function dni(v){
    const raw=String(v||'');
    const m=raw.match(/(?:^|\D)(\d{8})(?:\D|$)/);
    const d=m?m[1]:raw.replace(/\D/g,'');
    return d.length>=8?d.slice(-8):d;
  }
  function sound(ok=true){
    try{
      if(typeof beep==='function'){beep();return;}
      const C=window.AudioContext||window.webkitAudioContext; if(!C)return;
      const ctx=new C(), o=ctx.createOscillator(), g=ctx.createGain();
      o.type='sine'; o.frequency.value=ok?880:220; g.gain.value=.08;
      o.connect(g); g.connect(ctx.destination); o.start(); setTimeout(()=>{try{o.stop();ctx.close();}catch(e){}},140);
    }catch(e){}
  }
  function st(kind,msg){
    const e=$('dniStatus'); if(!e)return;
    e.className=(kind==='ok'?'scan-ok mt-2 flash':kind==='bad'?'scan-bad mt-2 flash':'mt-2 field-help');
    e.innerHTML=msg;
  }
  function render(){
    const q=$('workerQueue'), h=$('dnisMasivos'); if(!q||!h)return;
    const arr=[...window.__dniQueueFinal.entries()];
    h.value=arr.map(x=>x[0]).join(',');
    if(!arr.length){q.innerHTML='<div class="text-muted small text-center">Aún no hay trabajadores detectados.</div>';return;}
    q.innerHTML=arr.map(([d,n])=>'<div class="queue-item"><div><b>'+d+'</b><br><span>'+String(n||'TRABAJADOR')+'</span></div><button type="button" class="btn btn-sm btn-outline-danger" onclick="window.__dniQueueFinal.delete(\''+d+'\');window.renderQueueFinal&&window.renderQueueFinal();">×</button></div>').join('');
  }
  window.renderQueueFinal=render;
  async function process(v, force=false){
    const input=$('dniTrab'); if(!input)return;
    const d=dni(v||input.value);
    if(d.length<8){ if(force)st('help','Escanee o digite DNI: al completar 8 dígitos se agregará al pre-registro con sonido.'); return; }
    input.value=d;
    if(!force && d===last)return;
    if(busy){clearTimeout(timer); timer=setTimeout(()=>process(d,true),90); return;}
    busy=true; last=d; st('ok','Buscando DNI <b>'+d+'</b>...');
    try{
      const r=await fetch('/api/trabajador/'+encodeURIComponent(d),{cache:'no-store',credentials:'same-origin'});
      let j={ok:false,msg:'Respuesta inválida'}; try{j=await r.json();}catch(e){}
      if(!j.ok){st('bad','✕ '+(j.msg||'DNI no encontrado en base trabajadores')+' <b>'+d+'</b>'); sound(false); input.select(); return;}
      const t=j.trabajador||{}, nombre=t.trabajador||t.nombres||t.nombre||'TRABAJADOR';
      if(!window.__dniQueueFinal.has(d)){window.__dniQueueFinal.set(d,nombre);}
      if(window.workerMap){ try{window.workerMap.set(d,nombre); window.renderQueue&&window.renderQueue();}catch(e){} }
      render(); st('ok','✓ Reconocido automáticamente: <b>'+nombre+'</b> · '+d); sound(true);
      setTimeout(()=>{input.value=''; last=''; input.focus();},160);
    }catch(e){st('bad','Error consultando trabajador. Revisa conexión/sesión.'); sound(false);}
    finally{busy=false;}
  }
  window.autoDetectarDniInline=function(el){clearTimeout(timer); timer=setTimeout(()=>process((el&&el.value)||($('dniTrab')&&$('dniTrab').value)||'',false),25);};
  document.addEventListener('input',e=>{if(e.target&&e.target.id==='dniTrab')window.autoDetectarDniInline(e.target);},true);
  document.addEventListener('keyup',e=>{if(e.target&&e.target.id==='dniTrab')window.autoDetectarDniInline(e.target);},true);
  document.addEventListener('paste',e=>{if(e.target&&e.target.id==='dniTrab')setTimeout(()=>process(e.target.value,true),60);},true);
  document.addEventListener('keydown',e=>{if(e.target&&e.target.id==='dniTrab'&&(e.key==='Enter'||e.key==='Tab')){process(e.target.value,true); if(e.key==='Enter')e.preventDefault();}},true);
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalRegistro'){setTimeout(()=>{const i=$('dniTrab'); if(i){i.focus(); if(dni(i.value).length>=8)process(i.value,true);}},80);}},true);
  document.addEventListener('submit',e=>{if(e.target&&e.target.id==='frmTrab'){render(); const h=$('dnisMasivos'), i=$('dniTrab'); if(h&&i&&dni(i.value).length===8){const d=dni(i.value); if(!h.value.includes(d)) h.value=(h.value?h.value+',':'')+d;}}},true);
  setInterval(()=>{const i=$('dniTrab'); if(i&&dni(i.value).length>=8)process(i.value,false);},220);

  async function loadQr(){
    if(window.Html5Qrcode)return true;
    return new Promise(res=>{const s=document.createElement('script');s.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js';s.onload=()=>res(!!window.Html5Qrcode);s.onerror=()=>res(false);document.head.appendChild(s);});
  }
  let qr=null;
  window.abrirScanner=async function(readerId,inputId){
    const box=$(readerId), input=$(inputId); if(!box||!input)return;
    box.style.display='block'; box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="position:absolute;right:8px;top:8px;z-index:9999;border-radius:999px" onclick="cerrarScannerActivo()">×</button><div class="p-2 text-success fw-bold">Abriendo cámara...</div>';
    if(location.protocol!=='https:' && location.hostname!=='localhost' && location.hostname!=='127.0.0.1'){box.innerHTML='<div class="scan-bad">La cámara requiere HTTPS. Abre Render con https:// y permite cámara.</div>'; sound(false); return;}
    if(!navigator.mediaDevices){box.innerHTML='<div class="scan-bad">Este navegador no permite cámara. Usa Chrome actualizado.</div>'; sound(false); return;}
    if(!(await loadQr())){box.innerHTML='<div class="scan-bad">No cargó librería del lector. Revisa internet/CDN.</div>'; sound(false); return;}
    try{ if(qr){await qr.stop().catch(()=>{}); await qr.clear().catch(()=>{});} }catch(e){}
    try{
      qr=new Html5Qrcode(readerId);
      await qr.start({facingMode:{ideal:'environment'}},{fps:12,qrbox:{width:240,height:160}},async decoded=>{
        const d=dni(decoded); input.value=d; input.dispatchEvent(new Event('input',{bubbles:true}));
        if(inputId==='dniTrab') await process(d,true); else sound(true);
        try{await qr.stop(); await qr.clear();}catch(e){} box.style.display='none'; qr=null;
      },()=>{});
    }catch(e){box.innerHTML='<div class="scan-bad">No se pudo activar cámara. Permite cámara en el candado del navegador.</div>'; sound(false);}
  };
})();
</script>

<script>
/* ===== PATCH 248 JS: evitar doble lista nativa y hacer slider PC/táctil real por posición ===== */
(function(){
 const $=id=>document.getElementById(id);
 const pad=n=>String(Number(n)||0).padStart(2,'0');
 function minToTime(m){m=Math.max(0,Math.min(1435,parseInt(m||0,10)));return pad(Math.floor(m/60))+':'+pad(Math.round(m%60/5)*5).replace('60','55');}
 function toMin(v){let p=String(v||'00:00').split(':'),h=parseInt(p[0]||0,10),m=parseInt(p[1]||0,10);return Math.max(0,Math.min(1435,(isNaN(h)?0:h)*60+(isNaN(m)?0:m)));}
 let active='horaInicioDefault';
 function setActive(id){active=id; ['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'].forEach(x=>{let e=$(x); if(e)e.classList.toggle('border-success',x===id);}); let s=$('timeSlider24'), v=$('touchClockValue'), e=$(id); if(e&&s){s.value=toMin(e.value); if(v)v.textContent=e.value;} let box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>b.classList.toggle('active',b.dataset.target===id));}
 function sync(){let hi=$('horaInicioDefault')?.value||'06:30',hf=$('horaFinDefault')?.value||'16:30',ri=$('refInicioDefault')?.value||'12:00',rf=$('refFinDefault')?.value||'13:00'; [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,val])=>{let e=$(id); if(e)e.value=val;});}
 function apply(v){let e=$(active); if(!e)return; let t=minToTime(v); e.value=t; let d=$('touchClockValue'); if(d)d.textContent=t; sync();}
 function bindSlider(){let s=$('timeSlider24'); if(!s||s.dataset.patch248==='1')return; s.dataset.patch248='1';
   const calc=ev=>{let r=s.getBoundingClientRect(); let x=(ev.touches&&ev.touches[0]?ev.touches[0].clientX:ev.clientX); if(typeof x==='number'&&r.width){let pct=Math.max(0,Math.min(1,(x-r.left)/r.width)); let m=Math.round((pct*1435)/5)*5; s.value=m; apply(m);}else apply(s.value);};
   ['pointerdown','pointermove','mousedown','mousemove','touchstart','touchmove','click','input','change'].forEach(name=>s.addEventListener(name,ev=>{if(name.includes('move') && !(ev.buttons||ev.touches)) return; ev.preventDefault?.(); calc(ev);},{passive:false}));
 }
 function bindHorario(){let pills=$('clockPickFields'); if(pills)[...pills.querySelectorAll('button')].forEach(b=>{b.onclick=e=>{e.preventDefault();setActive(b.dataset.target);}; b.onpointerdown=e=>setActive(b.dataset.target);}); ['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'].forEach(id=>{let e=$(id); if(e){e.readOnly=true;e.onclick=()=>setActive(id);e.onpointerdown=()=>setActive(id);e.onfocus=()=>setActive(id);}}); bindSlider(); setActive(active); sync();}
 function fixModalLabor(){['modalActividad','modalLaborInput','modalConsumidor'].forEach(id=>{let e=$(id); if(e){e.removeAttribute('list'); e.setAttribute('autocomplete','off');}}); ['modal_actividad_list','modal_labor_list','modal_consumidor_list'].forEach(id=>{let d=$(id); if(d)d.innerHTML='';});}
 document.addEventListener('shown.bs.modal',ev=>{if(ev.target&&ev.target.id==='modalHora')setTimeout(bindHorario,80); if(ev.target&&ev.target.id==='modalLabor')setTimeout(fixModalLabor,30);});
 document.addEventListener('DOMContentLoaded',()=>{setTimeout(()=>{bindHorario();fixModalLabor();},120);});
 window.setCampoHorario=setActive; window.aplicarHorarioRegistro=sync;
})();
</script>

<script>
/* PATCH URGENTE: validación avance solo donde existe modalAvance, sin usar variables Jinja fuera de contexto */
(function(){
  const $=id=>document.getElementById(id);
  function onlyDni(v){const d=String(v||'').replace(/\D/g,'');return d.length>=8?d.slice(-8):d;}
  function setSt(ok,msg){const e=$('avanceTrabStatus'); if(!e)return; e.className=ok?'scan-ok mt-2':'scan-bad mt-2'; e.innerHTML=msg;}
  function hojaIdFromForm(){const f=$('frmAvance'); const m=f && String(f.getAttribute('action')||'').match(/\/hoja\/(\d+)\//); return m?m[1]:'';}
  async function validarTrabAvance(){
    const inp=$('dniAvance'), labor=$('avanceLaborId'); if(!inp||!labor)return;
    const dni=onlyDni(inp.value); if(dni.length<8){setSt(false,'Digite o escanee DNI de 8 dígitos.'); return;}
    inp.value=dni; const hoja=hojaIdFromForm();
    if(!hoja || !labor.value){setSt(false,'Primero seleccione una labor.'); return;}
    try{
      const r=await fetch('/api/trabajador-labor/'+encodeURIComponent(hoja)+'/'+encodeURIComponent(labor.value)+'/'+encodeURIComponent(dni),{cache:'no-store',credentials:'same-origin'});
      const j=await r.json();
      if(j.ok){setSt(true,'✓ Trabajador registrado en esta labor: <b>'+j.trabajador.trabajador+'</b>');}
      else{setSt(false,'✕ '+(j.msg||'El trabajador no está registrado en Trabajadores de esta labor.'));}
    }catch(e){setSt(false,'No se pudo validar trabajador.');}
  }
  function detectarCantidad(){
    const c=$('codigoAvance'), q=$('cantidadAvance'), box=$('cantidadDetectada'); if(!c||!q)return;
    const txt=String(c.value||''); const m=txt.match(/(\d+(?:[\.,]\d+)?)(?!.*\d)/);
    if(m){const val=m[1].replace(',','.'); q.value=Number(val).toFixed(2); if(box){box.style.display='block'; box.innerHTML='Cantidad detectada: <b>'+q.value+'</b>';}}
  }
  document.addEventListener('input',e=>{if(e.target&&e.target.id==='dniAvance')validarTrabAvance(); if(e.target&&e.target.id==='codigoAvance')detectarCantidad();},true);
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalAvance')setTimeout(validarTrabAvance,80);},true);
})();
</script>


<script>
/* === PATCH 249 FINAL JS: flecha acciones + horas deslizables/digitables === */
(function(){
  const $=id=>document.getElementById(id);
  const ids=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  const pad=n=>String(Number(n)||0).padStart(2,'0');
  function toMin(v){let m=String(v||'00:00').match(/^(\d{1,2}):(\d{2})$/); if(!m)return 0; let h=Math.max(0,Math.min(23,parseInt(m[1]||0,10))); let mi=Math.max(0,Math.min(59,parseInt(m[2]||0,10))); return h*60+mi;}
  function minToTime(m){m=((parseInt(m||0,10)%1440)+1440)%1440; return pad(Math.floor(m/60))+':'+pad(m%60);}
  function normTime(v){let s=String(v||'').replace(/[^0-9:]/g,''); if(/^\d{3,4}$/.test(s)){s=s.padStart(4,'0');s=s.slice(0,2)+':'+s.slice(2);} let m=s.match(/^(\d{1,2}):(\d{1,2})$/); if(!m)return null; let h=Math.min(23,Math.max(0,parseInt(m[1],10)||0)); let mi=Math.min(59,Math.max(0,parseInt(m[2],10)||0)); return pad(h)+':'+pad(mi);}
  function syncPreview(){try{ if(window.aplicarHorarioRegistro) window.aplicarHorarioRegistro(); }catch(e){} }
  window.toggleCardMenu=function(el){document.querySelectorAll('.card-menu.show').forEach(m=>{if(m!==el.nextElementSibling)m.classList.remove('show')}); const m=el.nextElementSibling; if(m)m.classList.toggle('show');};
  document.addEventListener('click',e=>{if(!e.target.closest('.card-menu')&&!e.target.closest('.card-action-chevron'))document.querySelectorAll('.card-menu.show').forEach(m=>m.classList.remove('show'));},true);
  function activate(id){try{ if(window.setCampoHorario) window.setCampoHorario(id); }catch(e){} ids.forEach(x=>{let e=$(x); if(e)e.classList.toggle('border-success',x===id);}); let v=$('touchClockValue'), s=$('timeSlider24'), el=$(id); if(el){if(v)v.textContent=el.value;if(s)s.value=toMin(el.value);} }
  function bindTimeBox(el){ if(!el||el.dataset.drag249==='1')return; el.dataset.drag249='1'; el.readOnly=false; el.classList.add('time-draggable'); el.setAttribute('inputmode','numeric'); el.setAttribute('placeholder','HH:MM');
    let startX=0,startVal=0,dragging=false,moved=false;
    const setVal=(mins)=>{el.value=minToTime(Math.round(mins/5)*5); activate(el.id); syncPreview();};
    el.addEventListener('pointerdown',ev=>{startX=ev.clientX;startVal=toMin(el.value);dragging=true;moved=false;activate(el.id); try{el.setPointerCapture(ev.pointerId)}catch(e){}}, {passive:true});
    el.addEventListener('pointermove',ev=>{if(!dragging)return; const dx=ev.clientX-startX; if(Math.abs(dx)>3)moved=true; setVal(startVal+Math.round(dx/6)*5); ev.preventDefault();}, {passive:false});
    el.addEventListener('pointerup',ev=>{dragging=false; if(moved){try{el.blur()}catch(e){}}}, {passive:true});
    el.addEventListener('click',()=>activate(el.id)); el.addEventListener('focus',()=>activate(el.id));
    el.addEventListener('change',()=>{const t=normTime(el.value); if(t)el.value=t; activate(el.id); syncPreview();});
    el.addEventListener('blur',()=>{const t=normTime(el.value); if(t)el.value=t; syncPreview();});
  }
  function bindAll(){ids.forEach(id=>bindTimeBox($(id))); let h=$('horarioCoherencia'); if(h&&!h.dataset.dragHelp){h.dataset.dragHelp='1';h.insertAdjacentHTML('beforeend','<div class="time-drag-help">Tip: toca un recuadro y arrastra a la derecha/izquierda para ajustar de 5 en 5 minutos. También puedes digitar HH:MM.</div>');}}
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalHora')setTimeout(bindAll,80);});
  document.addEventListener('DOMContentLoaded',()=>setTimeout(bindAll,150));
})();
</script>


<script>
/* === PATCH 250 OMAR FINAL: horario aislado, scanner con X, avance no se preregistra con solo DNI === */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  const pad=n=>String(Number(n)||0).padStart(2,'0');
  const ids=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  function normTime(v){let s=String(v||'').trim().replace(/[^0-9:]/g,''); if(/^\d{3,4}$/.test(s)){s=s.padStart(4,'0');s=s.slice(0,2)+':'+s.slice(2)} let m=s.match(/^(\d{1,2}):(\d{1,2})$/); if(!m)return null; let h=Math.max(0,Math.min(23,parseInt(m[1]||0,10))); let mi=Math.max(0,Math.min(59,parseInt(m[2]||0,10))); return pad(h)+':'+pad(Math.round(mi/5)*5).replace('60','55');}
  function toMin(v){let t=normTime(v)||'00:00',p=t.split(':');return parseInt(p[0],10)*60+parseInt(p[1],10)}
  function minToTime(m){m=((Math.round(Number(m||0)/5)*5)%1440+1440)%1440;return pad(Math.floor(m/60))+':'+pad(m%60)}
  function cloneClean(id){const e=$(id); if(!e||!e.parentNode)return null; const c=e.cloneNode(true); e.parentNode.replaceChild(c,e); return c;}
  function installHorarioOmar(){
    const modal=$('modalHora'); if(!modal)return;
    ids.forEach(cloneClean); cloneClean('timeSlider24'); cloneClean('clockPickFields');
    let active='horaInicioDefault';
    const inp=()=>$(active)||$('horaInicioDefault');
    function sync(){
      const hi=$('horaInicioDefault')?.value||'06:30', hf=$('horaFinDefault')?.value||'16:30', ri=$('refInicioDefault')?.value||'12:00', rf=$('refFinDefault')?.value||'13:00';
      [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,v])=>{const e=$(id); if(e)e.value=v;});
      let a=toMin(hi),b=toMin(hf); if(b<=a)b+=1440; let c=toMin(ri),d=toMin(rf); if(d<=c)d+=1440; if(b>1440&&c<a){c+=1440;d+=1440;} const ref=Math.max(0,Math.min(b,d)-Math.max(a,c)); const total=Math.max(0,(b-a)-ref)/60;
      const h=$('horasTrab'); if(h)h.value=total.toFixed(2);
      const txt=$('horarioActivoTxt'); if(txt)txt.innerHTML='<b>Horario activo:</b> '+hi+' - '+hf+' / Refrigerio '+ri+' - '+rf+' / H.Normal '+total.toFixed(2)+'.';
    }
    function paint(){const e=inp(),sl=$('timeSlider24'),tv=$('touchClockValue'); if(e&&sl)sl.value=toMin(e.value); if(e&&tv)tv.textContent=e.value; const box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>b.classList.toggle('active',b.dataset.target===active)); ids.forEach(id=>$(id)?.classList.toggle('border-success',id===active));}
    function setActive(id){if(ids.includes(id))active=id; paint();}
    function setVal(v){const e=inp(); if(!e)return; e.value=minToTime(v); sync(); paint();}
    ids.forEach(id=>{const e=$(id); if(!e)return; e.readOnly=false; e.classList.add('time-draggable'); e.setAttribute('inputmode','numeric');
      ['pointerdown','mousedown','touchstart','click','focus'].forEach(ev=>e.addEventListener(ev,evt=>{setActive(id); evt.stopPropagation();},true));
      e.addEventListener('change',()=>{const t=normTime(e.value); if(t)e.value=t; sync(); paint();}); e.addEventListener('blur',()=>{const t=normTime(e.value); if(t)e.value=t; sync(); paint();});
      let sx=0,sv=0,drag=false; e.addEventListener('pointerdown',ev=>{sx=ev.clientX;sv=toMin(e.value);drag=true;setActive(id);try{e.setPointerCapture(ev.pointerId)}catch(_){}}); e.addEventListener('pointermove',ev=>{if(!drag)return; const dx=ev.clientX-sx; if(Math.abs(dx)>2){setVal(sv+Math.round(dx/6)*5); ev.preventDefault();}}, {passive:false}); e.addEventListener('pointerup',()=>{drag=false;});
    });
    const box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(btn=>{['pointerdown','click','touchstart'].forEach(ev=>btn.addEventListener(ev,e=>{e.preventDefault();e.stopPropagation();setActive(btn.dataset.target);},true));});
    const sl=$('timeSlider24'); if(sl){const calc=ev=>{const r=sl.getBoundingClientRect(); const x=(ev.touches&&ev.touches[0])?ev.touches[0].clientX:ev.clientX; if(typeof x==='number'&&r.width){return Math.round(Math.max(0,Math.min(1,(x-r.left)/r.width))*1435/5)*5;} return Number(sl.value||0);}; ['pointerdown','pointermove','touchstart','touchmove','click','input','change'].forEach(ev=>sl.addEventListener(ev,e=>{if(ev.includes('move')&&!(e.buttons||e.touches))return; e.preventDefault&&e.preventDefault(); e.stopPropagation&&e.stopPropagation(); setVal(calc(e));},{passive:false,capture:true}));}
    setActive(active); sync();
  }
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalHora')setTimeout(installHorarioOmar,260);},true);
  document.addEventListener('DOMContentLoaded',()=>setTimeout(installHorarioOmar,500));

  window.cerrarScannerActivo=async function(){try{if(window.__qr250){await window.__qr250.stop().catch(()=>{});await window.__qr250.clear().catch(()=>{});}}catch(e){} document.querySelectorAll('[id^="reader"]').forEach(b=>{b.style.display='none';b.innerHTML='';});};
  const oldOpen=window.abrirScanner;
  window.abrirScanner=async function(readerId,inputId){
    const box=$(readerId), input=$(inputId); if(!box||!input)return;
    box.style.display='block'; box.style.position='relative'; box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="position:absolute;right:8px;top:8px;z-index:9999;border-radius:999px" onclick="cerrarScannerActivo()">×</button><div class="p-2 text-success fw-bold">Abriendo cámara...</div>';
    try{
      if(!window.Html5Qrcode){await new Promise((res,rej)=>{const s=document.createElement('script');s.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js';s.onload=res;s.onerror=rej;document.head.appendChild(s);});}
      await window.cerrarScannerActivo();
      box.style.display='block'; box.style.position='relative'; box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="position:absolute;right:8px;top:8px;z-index:9999;border-radius:999px">×</button>';
      box.querySelector('button').onclick=window.cerrarScannerActivo;
      window.__qr250=new Html5Qrcode(readerId);
      const cams=await Html5Qrcode.getCameras().catch(()=>[]); const camera=cams&&cams.length?{deviceId:{exact:cams[cams.length-1].id}}:{facingMode:{ideal:'environment'}};
      await window.__qr250.start(camera,{fps:12,qrbox:{width:240,height:160}},async decoded=>{let val=String(decoded||'').trim(); if(inputId==='dniTrab'||inputId==='dniAvance')val=String(val).replace(/\D/g,'').slice(-8); input.value=val; input.dispatchEvent(new Event('input',{bubbles:true})); input.dispatchEvent(new Event('change',{bubbles:true})); await window.cerrarScannerActivo();},()=>{});
    }catch(e){box.innerHTML='<button type="button" class="btn btn-sm btn-danger" style="float:right" onclick="cerrarScannerActivo()">×</button><div class="scan-bad mt-2">No se pudo activar cámara. Permite cámara en el candado del navegador o usa Chrome/Edge con HTTPS.</div>';}
  };

  let avanceAddTimer=null;
  document.addEventListener('input',e=>{
    if(e.target&&e.target.id==='dniAvance')setTimeout(()=>{const c=$('codigoAvance'); if(c&&String(e.target.value||'').replace(/\D/g,'').length>=8)c.focus();},250);
    if(e.target&&e.target.id==='codigoAvance'){clearTimeout(avanceAddTimer); avanceAddTimer=setTimeout(()=>{const ev=new KeyboardEvent('keydown',{key:'Enter',bubbles:true}); $('codigoAvance')?.dispatchEvent(ev);},500);}
  },true);
})();
</script>

<script>
/* === PATCH 251 OMAR: CANTIDAD + PRE-REGISTRO FINAL AVANCE + EDITAR HORARIO === */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  const dni=v=>{const d=String(v||'').replace(/\D/g,'');return d.length>=8?d.slice(-8):d;};
  const num=v=>{let m=String(v||'').replace(',','.').match(/(\d+(?:\.\d+)?)/);return m?Number(m[1]||0):0;};
  const key=()=>{const f=$('frmAvance');return 'avance_pre_'+(f?f.action:'')+'_'+($('avanceLaborId')?.value||'');};
  let avancePre=[];
  function load(){try{avancePre=JSON.parse(sessionStorage.getItem(key())||'[]')||[];}catch(e){avancePre=[];}}
  function save(){try{sessionStorage.setItem(key(),JSON.stringify(avancePre));}catch(e){} const h=$('avancePreJson'); if(h)h.value=JSON.stringify(avancePre);}
  function render(){const q=$('avanceQueue'); if(!q)return; save(); if(!avancePre.length){q.innerHTML='<div class="text-muted small text-center">Aún no hay avances pre-registrados.</div>';return;} q.innerHTML=avancePre.map((x,i)=>'<div class="queue-item"><div><b>'+x.dni+'</b><br><span>Cantidad: '+Number(x.cantidad||0).toFixed(2)+' · '+(x.metodo||'QR/CÓDIGO')+'</span></div><button type="button" class="btn btn-sm btn-outline-danger" data-del-avance="'+i+'">×</button></div>').join(''); q.querySelectorAll('[data-del-avance]').forEach(b=>b.onclick=()=>{avancePre.splice(Number(b.dataset.delAvance),1);render();});}
  function setMsg(ok,msg){const e=$('cantidadDetectada'); if(!e)return; e.style.display='block'; e.className=ok?'scan-ok mt-2':'scan-bad mt-2'; e.innerHTML=msg;}
  function addPre(){const d=dni($('dniAvance')?.value); const c=Number($('cantidadAvance')?.value||0); const labor=$('avanceLaborId')?.value||''; if(d.length!==8){setMsg(false,'Primero registre/escanee el DNI del trabajador.');return false;} if(!labor){setMsg(false,'Primero seleccione una labor.');return false;} if(!(c>0)){setMsg(false,'Ingrese una cantidad mayor a cero.');return false;} const met=$('frmAvance')?.querySelector('select[name="metodo"]')?.value||'QR/CÓDIGO'; const exists=avancePre.find(x=>x.dni===d); if(exists){exists.cantidad=c;exists.metodo=met;} else {avancePre.push({dni:d,cantidad:c,a_noct:0,metodo:met});} render(); setMsg(true,'✓ Pre-registro agregado: <b>'+d+'</b> · Cantidad <b>'+c.toFixed(2)+'</b>. Falta GUARDAR AVANCE FINAL.'); try{beep();}catch(e){} const ca=$('codigoAvance'); if(ca)ca.value=''; const da=$('dniAvance'); if(da)da.focus(); return true;}
  function installAvance(){const m=$('modalAvance'); if(!m)return; load(); render(); const cod=$('codigoAvance'), cant=$('cantidadAvance'); if(cod&&!cod.dataset.pre251){cod.dataset.pre251='1'; cod.addEventListener('input',()=>{const v=num(cod.value); if(v>0&&cant) {cant.value=v.toFixed(2); setMsg(true,'Cantidad detectada: <b>'+v.toFixed(2)+'</b>');}},true); cod.addEventListener('keydown',e=>{if(e.key==='Enter'||e.key==='Tab'){e.preventDefault(); addPre();}},true); cod.addEventListener('change',()=>{const v=num(cod.value); if(v>0&&cant) cant.value=v.toFixed(2); addPre();},true);} const form=$('frmAvance'); if(form&&!form.dataset.pre251){form.dataset.pre251='1'; form.addEventListener('submit',e=>{if(!avancePre.length){addPre();} if(!avancePre.length){e.preventDefault(); return false;} save(); sessionStorage.removeItem(key());},true);} }
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalAvance')setTimeout(installAvance,80);},true);
  document.addEventListener('DOMContentLoaded',()=>setTimeout(installAvance,300));
  window.abrirEditarAvance=function(id,cantidad){const f=$('frmEditAvance'); if(!f)return; f.action='/lectura/'+id+'/editar'; $('editCantAvance').value=Number(cantidad||0).toFixed(2); if($('editNoctAvance')) $('editNoctAvance').value='0.00'; new bootstrap.Modal($('modalEditAvance')).show();};
})();
</script>
<style>
.avance-grid.cantidad-only{grid-template-columns:1fr 1fr 32px!important}.avance-grid.cantidad-only label{font-size:8px!important;font-weight:900;color:#55745a}.avance-grid.cantidad-only .mini-badge{height:26px!important}.card-menu button{cursor:pointer}.card-menu button:hover{background:#eef8ef!important}.modal-title:has(+ .btn-close){}
</style>

<script>
/* PATCH 253: botón visible para apagar cámara + scanner robusto */
(function(){
  'use strict';
  const $=id=>document.getElementById(id);
  function closeBtn(){return '<button type="button" class="scanner-close-x" aria-label="Apagar cámara" title="Apagar cámara">×</button>';}
  function ponerX(box){
    if(!box) return;
    box.style.position='relative';
    if(!box.querySelector('.scanner-close-x')) box.insertAdjacentHTML('afterbegin', closeBtn());
    const b=box.querySelector('.scanner-close-x');
    if(b){ b.onclick=function(ev){ev.preventDefault();ev.stopPropagation(); window.cerrarScannerActivo&&window.cerrarScannerActivo();}; }
  }
  async function detener(obj){try{if(obj){await obj.stop?.().catch(()=>{}); await obj.clear?.().catch(()=>{});}}catch(e){}}
  window.cerrarScannerActivo=async function(){
    await detener(window.__scannerActivo253); window.__scannerActivo253=null;
    await detener(window.__qr250); window.__qr250=null;
    document.querySelectorAll('[id^="reader"]').forEach(el=>{el.style.display='none'; el.innerHTML='';});
    try{document.querySelectorAll('video').forEach(v=>{const st=v.srcObject; if(st&&st.getTracks)st.getTracks().forEach(t=>t.stop());});}catch(e){}
  };
  function loadQr(){return new Promise(res=>{if(window.Html5Qrcode)return res(true); const s=document.createElement('script'); s.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js'; s.onload=()=>res(!!window.Html5Qrcode); s.onerror=()=>res(false); document.head.appendChild(s);});}
  window.abrirScanner=async function(readerId,inputId){
    const box=$(readerId), input=$(inputId); if(!box||!input)return;
    await window.cerrarScannerActivo();
    box.style.display='block'; box.innerHTML='<div class="p-2 text-success fw-bold">Abriendo cámara...</div>'; ponerX(box);
    if(location.protocol!=='https:' && location.hostname!=='localhost' && location.hostname!=='127.0.0.1'){
      box.innerHTML='<div class="scan-bad mt-2">La cámara requiere HTTPS. Abre Render con https:// y permite cámara.</div>'; ponerX(box); return;
    }
    if(!(await loadQr())){box.innerHTML='<div class="scan-bad mt-2">No cargó la librería del lector.</div>'; ponerX(box); return;}
    try{
      const scanner=new Html5Qrcode(readerId); window.__scannerActivo253=scanner;
      const cams=await Html5Qrcode.getCameras().catch(()=>[]);
      const camera=(cams&&cams.length)?{deviceId:{exact:cams[cams.length-1].id}}:{facingMode:{ideal:'environment'}};
      await scanner.start(camera,{fps:12,qrbox:{width:260,height:180}},async decoded=>{
        let val=String(decoded||'').trim();
        if(inputId==='dniTrab'||inputId==='dniAvance') val=val.replace(/\D/g,'').slice(-8);
        input.value=val; input.dispatchEvent(new Event('input',{bubbles:true})); input.dispatchEvent(new Event('change',{bubbles:true}));
        await window.cerrarScannerActivo();
      },()=>{});
      ponerX(box);
      const mo=new MutationObserver(()=>ponerX(box)); mo.observe(box,{childList:true,subtree:false}); setTimeout(()=>mo.disconnect(),15000);
    }catch(e){box.innerHTML='<div class="scan-bad mt-2">No se pudo activar cámara. Permite cámara en el candado del navegador.</div>'; ponerX(box);}
  };
})();
</script>
</body></html>
"""

def render_page(body, title="Tareo Móvil", **ctx):
    return render_template_string(BASE_HTML, body=render_template_string(body, **ctx), title=title)


def get_actividades_maestras(limit=50000):
    """Devuelve actividades únicas.
    Antes estaba limitado a 5000, por eso con cargas de 20 mil filas el modal no veía todo.
    Además se usa DISTINCT para que no repita la misma ACTIVIDAD/LABOR/CONSUMIDOR.
    """
    try:
        sql = """SELECT DISTINCT
                    COALESCE(cod_actividad,'') AS cod_actividad,
                    COALESCE(desc_actividad,'') AS desc_actividad,
                    COALESCE(cod_labor,'') AS cod_labor,
                    COALESCE(desc_labor,'') AS desc_labor,
                    COALESCE(cod_consumidor,'') AS cod_consumidor,
                    COALESCE(desc_consumidor,'') AS desc_consumidor
                 FROM actividades_maestras
                 WHERE COALESCE(estado,'ACTIVO')='ACTIVO'
                 ORDER BY desc_actividad, desc_labor, desc_consumidor
                 LIMIT ?"""
        rows = rows_to_dict(execute(sql, (limit,), fetchall=True))
    except Exception:
        rows = []
    return rows

def js_master_options(rows):
    import json
    return json.dumps(rows, ensure_ascii=False)

# ========================= AUTH + HOME =========================

@app.route('/inicio')
def inicio():
    body = """
    <div class="phone-wrap"><div class="splash-card">
      <div class="splash-logo"><i class="bi bi-clipboard2-data"></i></div>
      <div class="splash-title">TAREO MOVIL</div>
      <a class="btn btn-light btn-sm mt-4 fw-bold text-success" href="{{url_for('login')}}">ENTRAR</a>
      <div class="splash-foot">P&A<br>v.1.0</div>
    </div></div>"""
    return render_page(body, title='Tareo Móvil')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario','').strip()
        password = request.form.get('password','')
        row = row_to_dict(execute('SELECT * FROM usuarios WHERE usuario=?', (usuario,), fetchone=True))
        if row and row.get('estado','ACTIVO') == 'ACTIVO' and check_password_hash(row['password_hash'], password):
            if request.form.get('modo') == 'admin' and row.get('rol') != 'admin':
                flash('Este acceso es solo para administradores.', 'danger'); return redirect(url_for('login'))
            session['usuario']=usuario; session['rol']=row.get('rol','operador'); session['nombres']=row.get('nombres') or usuario
            return redirect(url_for('home'))
        flash('Usuario/clave incorrecta o usuario inactivo.', 'danger')
    body = """
    <div class="phone-wrap">
      <div class="green-hero" style="min-height:245px;border-radius:0 0 22px 22px">
        <div class="green-top"><span><i class="bi bi-headset"></i> Soporte</span><span><i class="bi bi-gear"></i> Config.</span></div>
        <div class="splash-logo" style="width:96px;height:96px;font-size:42px;margin:14px auto 6px"><i class="bi bi-clipboard2-data"></i></div>
        <div class="splash-title">TAREO MOVIL</div><div class="login-name">INICIAR SESIÓN</div>
      </div>
      <form method="post" class="login-form">
        <div class="floating-card">
          <div class="role-toggle mb-2"><label><input type="radio" name="modo" value="usuario" checked><span>USUARIO</span></label><label><input type="radio" name="modo" value="admin"><span>ADMINISTRADOR</span></label></div>
          <input class="form-control white-input mb-2" name="usuario" required autofocus placeholder="Usuario">
          <input class="form-control white-input mb-3" name="password" type="password" required placeholder="Clave">
          <button class="btn btn-green w-100"><i class="bi bi-box-arrow-in-right me-1"></i> INGRESAR</button>
          <div class="text-center small mt-2 text-muted">Admin demo: admin / admin123</div>
        </div>
      </form>
      <div class="d-flex justify-content-center gap-3 mt-4">
        <div class="tile"><i class="bi bi-list-check"></i>TAREO</div><div class="tile"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</div>
      </div>
      <div class="leaf"></div>
      <div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Sincronizar Tablas Maestras<br>Actualizado hasta: {{ now }}</div><a class="bottom-out"><i class="bi bi-box-arrow-right"></i></a>
    </div>"""
    return render_page(body, title='Login Tareo Móvil', now=now_str())

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/')
def home():
    if not session.get('usuario'):
        return redirect(url_for('inicio'))
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 8', fetchall=True))
    body = """
    <div class="desktop-grid">
      <div class="phone-wrap">
        <div class="green-hero" style="min-height:220px">
          <div class="green-top"><a class="text-white text-decoration-none" href="{{url_for('soporte')}}"><i class="bi bi-headset"></i> Soporte</a>{% if session.get('rol')=='admin' %}<a class="text-white text-decoration-none" href="{{url_for('configuraciones')}}"><i class="bi bi-gear"></i> Config.</a>{% else %}<span></span>{% endif %}</div>
          <div class="avatar"><i class="bi bi-person-circle"></i></div><div class="login-name">{{ session.get('nombres','USUARIO') }}</div>
          <div class="white-input mt-3"></div>
        </div>
        <div class="top-actions">
          <a class="tile text-decoration-none" href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>TAREO</a>
          <a class="tile text-decoration-none" href="{{url_for('reportes')}}"><i class="bi bi-file-earmark-bar-graph"></i>REPORTES<br>TAREO</a>
          <a class="tile text-decoration-none" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-repeat"></i>SINC.</a>
        </div>
        <div class="leaf"></div>
        <div class="bottom-sync"><i class="bi bi-arrow-repeat"></i> Sincronizar Tablas Maestras<br>Actualizado hasta: {{ now }}</div><a href="{{url_for('logout')}}" class="bottom-out"><i class="bi bi-box-arrow-right"></i></a>
      </div>
      <div class="desk-panel">
        <h1 class="header-title">TAREO MÓVIL – GRUPO DE COSECHA</h1>
        <div class="card-pro p-4 mb-3"><div class="d-flex justify-content-between align-items-center"><div><h4 class="fw-bold text-success mb-1">Hojas recientes</h4><div class="text-muted small">Crea una hoja y registra labores, trabajadores y avances.</div></div><a class="btn btn-green" href="{{url_for('crear_hoja')}}"><i class="bi bi-plus-lg"></i> Crear hoja</a></div></div>
        <div class="card-pro p-3"><div class="table-responsive"><table class="table list-table"><thead><tr><th>Fecha</th><th>Grupo</th><th>Subgrupo</th><th>Labor</th><th>Responsable</th><th>Estado</th><th></th></tr></thead><tbody>{% for h in hojas %}<tr><td>{{h.fecha}}</td><td>{{h.grupo}}</td><td>{{h.subgrupo}}</td><td>{{h.labor}}</td><td>{{h.responsable}}</td><td><span class="status-pill">{{h.estado}}</span></td><td><a class="btn btn-sm btn-green" href="{{url_for('detalle_hoja', hoja_id=h.id)}}">Abrir</a></td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted py-4">Sin hojas creadas.</td></tr>{% endfor %}</tbody></table></div></div>
      </div>
    </div>"""
    return render_page(body, hojas=hojas, now=now_str())

# ========================= HOJAS TAREO =========================
@app.route('/hojas')
@login_required
def hojas_tareo():
    hojas = rows_to_dict(execute('SELECT * FROM hojas_tareo ORDER BY fecha DESC, id DESC LIMIT 50', fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad tareo-list-page">
      <div class="page-card">
        <div class="green-hero tareo-hero" style="border-radius:0 0 12px 12px">
          <div class="green-top"><span>v.1.0</span><a class="text-white text-decoration-none" href="{{url_for('home')}}"><i class="bi bi-house"></i></a></div>
          <i class="bi bi-list-check" style="font-size:34px;margin-top:14px"></i><div class="login-name mt-1">TAREOS</div>
        </div>
        <div class="toolstrip tareo-toolbar">
          <a class="btn-plus-fab" title="Crear hoja" href="{{url_for('crear_hoja')}}"><i class="bi bi-list-task"></i><i class="bi bi-plus-circle-fill"></i></a>
          <a title="Plantilla Excel" href="{{url_for('plantilla_trabajadores')}}"><i class="bi bi-file-earmark-excel"></i></a>
          <a title="Sincronizar" href="{{url_for('sincronizacion')}}"><i class="bi bi-arrow-clockwise"></i></a>
        </div>
        {% for h in hojas %}
          <div class="swipe-wrap">
            <div class="swipe-actions"><a class="act-edit" href="{{url_for('editar_hoja', hoja_id=h.id)}}">MODIFICAR</a><a class="act-send" href="{{url_for('enviar_hoja', hoja_id=h.id)}}">ENVIAR</a><a class="act-del" href="{{url_for('eliminar_hoja', hoja_id=h.id)}}" onclick="return confirm('¿Eliminar esta hoja?')">ELIMINAR</a></div>
            <a class="text-decoration-none" href="{{url_for('detalle_hoja', hoja_id=h.id, tab='labores')}}"><div class="worker-card">
              <span class="person-dot" style="border-radius:8px"><i class="bi bi-clipboard2-check"></i></span>
              <div class="worker-title"><div>RESPONSABLE<br><b>{{h.responsable}}</b></div><div class="text-end">PRESUPUESTO<br><b>{{h.tipo_tareo or 'JORNAL'}}</b></div></div>
              <div class="worker-grid"><div><label>SUCURSAL</label><div class="small-value">{{h.grupo}}</div></div><div><label>PLANILLA</label><div class="small-value">AGR. PACKING</div></div><div><label>DOCUMENTO</label><div class="small-value">{{h.id}}</div></div></div>
              <div class="small-label mt-2">ZONA CONSUMIDOR</div><div class="small-value">{{h.subgrupo}}</div>
              <div class="worker-grid"><div><label>FECHA</label><div class="small-value">{{h.fecha}}</div></div><div><label>ESTADO</label><div class="mini-badge bg-y">{{h.estado}}</div></div><div class="text-end"><i class="bi bi-chevron-left text-success"></i></div></div>
            </div></a>
          </div>
        {% else %}<div class="worker-card text-center text-muted">No hay hojas. Presiona <b>+</b> para crear.</div>{% endfor %}
        <div class="leaf"></div>
        <div class="bottom-nav"><a href="{{url_for('hojas_tareo')}}"><i class="bi bi-list-check"></i>Listado de Tareos</a><a href="{{url_for('home')}}"><i class="bi bi-file-text"></i>Detalle</a></div>
      </div>
    </div>"""
    return render_page(body, hojas=hojas)

@app.route('/hojas/crear', methods=['GET','POST'])
@login_required
def crear_hoja():
    if request.method == 'POST':
        fecha = request.form.get('fecha') or today_str()
        grupo = limpiar_texto(request.form.get('actividad') or request.form.get('grupo'))
        subgrupo = limpiar_texto(request.form.get('labor') or request.form.get('subgrupo'))
        labor = limpiar_texto(request.form.get('consumidor') or request.form.get('consumidor_desc') or request.form.get('labor_consumidor') or '')
        responsable = limpiar_texto(request.form.get('responsable'))
        turno = limpiar_texto(request.form.get('turno') or 'DIA')
        tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or 'JORNAL')
        if turno not in ('DIA','NOCHE'): turno = 'DIA'
        if tipo_tareo not in ('JORNAL','RENDIMIENTO'): tipo_tareo = 'JORNAL'
        if not grupo or not subgrupo or not responsable:
            flash('Debe seleccionar Actividad y Labor, además de responsable.', 'danger')
            return redirect(url_for('crear_hoja'))
        execute('INSERT INTO hojas_tareo(fecha,grupo,subgrupo,labor,responsable,turno,tipo_tareo,estado,creado_por,creado_en) VALUES(?,?,?,?,?,?,?,?,?,?)',
                (fecha,grupo,subgrupo,labor,responsable,turno,tipo_tareo,'ABIERTA',session.get('usuario'),now_str()), commit=True)
        hid = scalar('SELECT MAX(id) AS id FROM hojas_tareo')
        execute('INSERT INTO hoja_labores(hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)',
                (hid,grupo,subgrupo,labor,turno,tipo_tareo,responsable,now_str(),session.get('usuario')), commit=True)
        return redirect(url_for('hojas_tareo'))
    maestros = get_actividades_maestras()
    body = """
    <div id="createHojaCompact" class="phone-wrap desktop-pad"><h2 class="header-title">CREAR HOJA DE TAREO</h2><div class="page-card">
      <div class="panel-green"><i class="bi bi-clipboard2-plus"></i><h4>NUEVA HOJA – FECHA, ACTIVIDAD, LABOR, CONSUMIDOR, RESPONSABLE, TURNO Y TIPO</h4></div>
      <form method="post" class="floating-card" style="margin:-24px 10px 12px">
        <label class="form-label">FECHA</label><input type="date" name="fecha" class="form-control mb-2 field-required" value="{{today}}" required>
        <label class="form-label">ACTIVIDAD</label><input id="actividadInput" name="actividad" class="form-control mb-2 field-required" list="actividad_list" placeholder="Digite primeras letras de la actividad" required autocomplete="off"><datalist id="actividad_list"></datalist>
        <label class="form-label">LABOR</label><input id="laborInput" name="labor" class="form-control mb-2 field-required" list="labor_list" placeholder="Seleccione labor según actividad" required autocomplete="off"><datalist id="labor_list"></datalist>
        <label class="form-label">CONSUMIDOR <span class="text-muted">(opcional)</span></label><input id="consumidorInput" name="consumidor" class="form-control mb-2" list="consumidor_list" placeholder="Consumidor / zona / campo"><datalist id="consumidor_list"></datalist>
        <label class="form-label">RESPONSABLE</label><input name="responsable" class="form-control mb-2 field-required" placeholder="APELLIDOS Y NOMBRES" required>
        <div class="row g-2 mb-3"><div class="col-6"><label class="form-label">TURNO</label><select name="turno" class="form-select field-required"><option>DIA</option><option>NOCHE</option></select></div><div class="col-6"><label class="form-label">TIPO</label><select name="tipo_tareo" class="form-select field-required"><option>JORNAL</option><option>RENDIMIENTO</option></select></div></div>
        <button class="btn btn-green w-100"><i class="bi bi-check-circle"></i> GUARDAR HOJA</button><a class="btn btn-outline-secondary w-100 mt-2" href="{{url_for('hojas_tareo')}}">VOLVER</a>
      </form></div></div>
      <script>
        const MAESTROS={{ maestros_json|safe }};
        const uniq=a=>[...new Set(a.filter(Boolean))].sort();
        function fillList(id, arr){const dl=document.getElementById(id); dl.innerHTML=''; arr.slice(0,300).forEach(v=>{const o=document.createElement('option'); o.value=v; dl.appendChild(o);});}
        function refreshActividad(){fillList('actividad_list', uniq(MAESTROS.map(x=>x.desc_actividad || x.cod_actividad)));}
        function refreshLabor(){const a=(actividadInput.value||'').toUpperCase(); const rows=MAESTROS.filter(x=>!a || String(x.desc_actividad||'').toUpperCase()===a || String(x.cod_actividad||'').toUpperCase()===a || String(x.desc_actividad||'').toUpperCase().includes(a)); fillList('labor_list', uniq(rows.map(x=>x.desc_labor || x.cod_labor))); refreshConsumidor();}
        function refreshConsumidor(){const a=(actividadInput.value||'').toUpperCase(), l=(laborInput.value||'').toUpperCase(); const rows=MAESTROS.filter(x=>(!a || String(x.desc_actividad||'').toUpperCase().includes(a)||String(x.cod_actividad||'').toUpperCase().includes(a)) && (!l || String(x.desc_labor||'').toUpperCase().includes(l)||String(x.cod_labor||'').toUpperCase().includes(l))); fillList('consumidor_list', uniq(rows.map(x=>x.desc_consumidor || x.cod_consumidor)));}
        actividadInput.addEventListener('input', refreshLabor); laborInput.addEventListener('input', refreshConsumidor); document.addEventListener('DOMContentLoaded',()=>{refreshActividad();refreshLabor();});
      </script>

    """
    return render_page(body, today=today_str(), maestros_json=js_master_options(maestros))

@app.route('/hoja/<int:hoja_id>')
@login_required
def detalle_hoja(hoja_id):
    tab = request.args.get('tab','labores')
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger'); return redirect(url_for('hojas_tareo'))
    labores = rows_to_dict(execute('SELECT * FROM hoja_labores WHERE hoja_id=? ORDER BY id DESC', (hoja_id,), fetchall=True))
    selected_labor_id = limpiar_texto(request.args.get('labor_id') or '', upper=False)
    selected_labor = None
    if selected_labor_id:
        selected_labor = row_to_dict(execute('SELECT * FROM hoja_labores WHERE id=? AND hoja_id=?', (selected_labor_id, hoja_id), fetchone=True))
        if not selected_labor:
            selected_labor_id = ''
    if tab in ('trabajadores','rendimiento') and not selected_labor_id:
        flash('Primero debes elegir una labor: toca/clic en una tarjeta de labor para entrar a Trabajadores o Rend./Avance.', 'danger')
        tab = 'labores'
    if selected_labor_id:
        tareos = rows_to_dict(execute('SELECT * FROM tareos WHERE hoja_id=? AND labor_id=? ORDER BY creado_en DESC LIMIT 100', (hoja_id, selected_labor_id), fetchall=True))
        lecturas = rows_to_dict(execute('SELECT * FROM lecturas_balde WHERE hoja_id=? AND labor_id=? ORDER BY fecha_hora DESC LIMIT 100', (hoja_id, selected_labor_id), fetchall=True))
    else:
        tareos = rows_to_dict(execute('SELECT * FROM tareos WHERE hoja_id=? ORDER BY creado_en DESC LIMIT 100', (hoja_id,), fetchall=True))
        lecturas = rows_to_dict(execute('SELECT * FROM lecturas_balde WHERE hoja_id=? ORDER BY fecha_hora DESC LIMIT 100', (hoja_id,), fetchall=True))
    registros = len(tareos); horas_total = sum(float(x.get('horas') or 0) for x in tareos); rend_total = sum(float(x.get('cantidad') or 0) for x in tareos)
    execute('UPDATE hojas_tareo SET registros=?, horas_total=?, rendimiento_total=? WHERE id=?', (registros, horas_total, rend_total, hoja_id), commit=True)
    body = """
    <div class="phone-wrap desktop-pad"><h2 class="header-title">TAREO MÓVIL – {{ 'DETALLE DE TRABAJADOR POR LABOR' if tab=='trabajadores' else ('DETALLE NÚMERO DE LECTURAS POR BALDE' if tab=='rendimiento' else 'GRUPO DE COSECHA') }}</h2>
      <div class="page-card">
        <div class="tab-main"><a class="{{'active' if tab=='labores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}">LABORES</a><a class="{{'active' if tab=='trabajadores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='trabajadores', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">TRABAJADORES</a><a class="{{'active' if tab=='rendimiento' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='rendimiento', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">REND./AVANCE</a></div>
        <div class="subtabs"><a class="{{'active' if tab=='labores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}">Labores</a><a class="{{'active' if tab=='trabajadores' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='trabajadores', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">Trab.por Labor</a><a class="{{'active' if tab=='rendimiento' else ''}}" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='rendimiento', labor_id=selected_labor_id) if selected_labor_id else url_for('detalle_hoja',hoja_id=h.id,tab='labores')}}" onclick="{% if not selected_labor_id %}alert('Primero toca/clic en una labor.');{% endif %}">Rend/Avance por Labor</a></div>
        <div class="panel-green"><i class="bi {{ 'bi-people-fill' if tab=='trabajadores' else ('bi-person-badge' if tab=='rendimiento' else 'bi-people') }}"></i><h4>{{ 'TRABAJADORES – QR / CÓDIGO BARRAS / DIGITACIÓN' if tab=='trabajadores' else ('PRODUCTIVIDAD POR TRABAJADOR – BALDE / QR / DIGITACIÓN' if tab=='rendimiento' else 'REGISTRO DE ACTIVIDAD, LABOR, CONSUMIDOR, TURNO Y TIPO') }}</h4></div>
        <div class="toolstrip">
          <button title="Crear labor/grupo/subgrupo" data-bs-toggle="modal" data-bs-target="#modalLabor"><i class="bi bi-list-check"></i></button>
          <button title="Buscar trabajador" data-bs-toggle="modal" data-bs-target="#modalBuscar"><i class="bi bi-search"></i></button>
          {% if tab=='trabajadores' %}
            <button title="Elegir horarios" data-bs-toggle="modal" data-bs-target="#modalHora"><i class="bi bi-clock"></i></button>
            <button title="Entrada / salida" data-bs-toggle="modal" data-bs-target="#modalHora"><i class="bi bi-box-arrow-in-right"></i></button>
            <button title="Registrar trabajador" data-bs-toggle="modal" data-bs-target="#modalRegistro"><i class="bi bi-person-plus"></i></button>
          {% elif tab=='rendimiento' %}
            <button title="Registrar avance" data-bs-toggle="modal" data-bs-target="#modalAvance"><i class="bi bi-upc-scan"></i></button>
            <button title="Refrescar" onclick="location.reload()"><i class="bi bi-arrow-clockwise"></i></button>
          {% else %}
            <button title="Copiar labor" data-bs-toggle="modal" data-bs-target="#modalCopiar"><i class="bi bi-files"></i></button>
            <button title="Refrescar" onclick="location.reload()"><i class="bi bi-arrow-clockwise"></i></button>
          {% endif %}
          <a href="{{url_for('hojas_tareo')}}" title="Volver"><i class="bi bi-chevron-left"></i></a>
        </div>
        <div class="info-bar"><div><i class="bi bi-calendar"></i> {{h.fecha}}</div><div><i class="bi bi-list"></i> {{registros}} Reg.</div><div><i class="bi bi-clock"></i> {{'%.2f'|format(horas_total)}} H.</div><div>A.Rend {{'%.2f'|format(rend_total)}}</div><span>⌄</span></div>
        {% if tab=='labores' %}
          {% for l in labores %}<a class="text-decoration-none" href="{{url_for('detalle_hoja',hoja_id=h.id,tab='trabajadores', labor_id=l.id)}}"><div class="worker-card labor-card-compact"><div class="worker-title"><div>ACTIVIDAD<br><b>{{l.grupo}}</b></div><div class="text-end">LABOR<br><b>{{l.subgrupo or 'SIN LABOR'}}</b></div></div><div class="mt-2"><span class="small-label">CONSUMIDOR</span> <b class="labor-main">{{l.labor or 'SIN CONSUMIDOR'}}</b><br><span class="small-label">RESPONSABLE</span> <b class="resp-main">{{l.responsable or h.responsable}}</b></div><div class="worker-grid mt-2"><div><div class="mini-badge {{'bg-y' if l.turno=='NOCHE' else 'bg-g'}}">{{l.turno}}</div></div><div><div class="mini-badge bg-y">{{l.tipo_tareo}}</div></div><div><div class="mini-badge bg-g">ACTIVA</div></div></div></div></a>{% else %}<div class="worker-card text-center text-muted">Presiona <b>+</b> para crear actividad, labor y consumidor.</div>{% endfor %}
        {% elif tab=='trabajadores' %}
          {% for r in tareos %}<div class="worker-card trabajador-card-ref {{'editable-tareo' if h.estado!='ENVIADA' else ''}}"><div class="worker-title"><div>TRABAJADOR<br><b>{{r.trabajador}}</b></div><div>NRO.DOCUMENTO<br><b>{{r.dni}}</b></div></div><div class="trabajador-grid-ref"><div><label>H.INICIO</label><div class="time-box">{{r.hora_inicio or ('22:00' if r.turno=='NOCHE' else '06:30')}}</div></div><div><label>H.FIN</label><div class="time-box">{{r.hora_fin or ('06:00' if r.turno=='NOCHE' else '16:30')}}</div></div><div><label>H.NORMAL</label><div class="metric-box">{{'%.2f'|format((r.horas or 0) - (r.horas_nocturnas or 0))}}</div></div><div><label>REF.INI</label><div class="time-box">{{r.ref_inicio or '12:00'}}</div></div><div><label>REF.FIN</label><div class="time-box">{{r.ref_fin or '13:00'}}</div></div><div><label>H.NOCTURNO</label><div class="metric-box">{{'%.2f'|format(r.horas_nocturnas or 0)}}</div></div><div><label>ESTADO</label><div class="mini-badge bg-g">FIN TOTAL</div></div></div>{% if h.estado!='ENVIADA' %}<div class="card-action-chevron" onclick="event.stopPropagation();toggleCardMenu(this)" title="Acciones"></div><div class="card-menu" onclick="event.stopPropagation()"><a href="{{url_for('editar_horas_tareo_form', tareo_id=r.id)}}" class="btn-edit-tareo" onclick="abrirEditarTareo('{{r.id}}','{{r.hora_inicio or ('22:00' if r.turno=='NOCHE' else '06:30')}}','{{r.hora_fin or ('06:00' if r.turno=='NOCHE' else '16:30')}}','{{r.ref_inicio or '12:00'}}','{{r.ref_fin or '13:00'}}'); return false;" data-id="{{r.id}}" data-hi="{{r.hora_inicio or ('22:00' if r.turno=='NOCHE' else '06:30')}}" data-hf="{{r.hora_fin or ('06:00' if r.turno=='NOCHE' else '16:30')}}" data-ri="{{r.ref_inicio or '12:00'}}" data-rf="{{r.ref_fin or '13:00'}}">Modificar</a><a class="danger" href="{{url_for('eliminar_tareo', tareo_id=r.id)}}" onclick="return confirm('¿Eliminar trabajador del tareo?')">Eliminar</a></div>{% endif %}</div>{% else %}<div class="worker-card text-center text-muted">Presiona el <b>hombresito +</b> para registrar trabajador por QR/código/digitación.</div>{% endfor %}
        {% else %}
          {% for l in lecturas %}<div class="worker-card avance-card-ref"><span class="person-dot"><i class="bi bi-person-circle"></i></span><div class="worker-title"><div>TRABAJADOR<br><b>{{l.trabajador}}</b></div><div>NRO.DOC.<br><b>{{l.dni}}</b></div></div><div class="small-label mt-1">HORA TOMA REGISTRO</div><div class="small-value">{{l.fecha_hora}} · {{l.metodo or 'DIGITACIÓN'}}</div><div class="worker-grid avance-grid cantidad-only"><div><label>CANTIDAD</label><div class="mini-badge bg-y">{{'%.2f'|format(l.a_diurno or 0)}}</div></div><div><label>UNIDAD</label><div class="mini-badge bg-y">BALDE</div></div><div></div></div>{% if h.estado!='ENVIADA' %}<div class="card-action-chevron" onclick="event.stopPropagation();toggleCardMenu(this)" title="Acciones"></div><div class="card-menu" onclick="event.stopPropagation()"><button type="button" onclick="abrirEditarAvance('{{l.id}}','{{l.a_diurno or 0}}')">Modificar</button><a class="danger" href="{{url_for('eliminar_lectura', lectura_id=l.id)}}" onclick="return confirm('¿Eliminar avance?')">Eliminar</a></div>{% else %}<i class="bi bi-lock text-muted"></i>{% endif %}</div>{% else %}<div class="worker-card text-center text-muted">Presiona el icono de escaneo para registrar avance por QR/código/digitación.</div>{% endfor %}
        {% endif %}<div class="leaf"></div>
      </div>
    </div>

    <div class="modal fade" id="modalLabor" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('guardar_labor_hoja', hoja_id=h.id, tab=tab)}}"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-plus-square"></i> Crear nueva labor</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small mb-2">Complete los datos y presione <b>CREAR LABOR</b>. Al cerrar con X no se guarda nada.</div><label class="form-label">ACTIVIDAD</label><input id="modalActividad" name="grupo" class="form-control mb-1" placeholder="Digite primeras letras de la actividad" required autocomplete="off"><datalist id="modal_actividad_list"></datalist><div id="modalMasterStatus" class="master-status">CARGANDO ACTIVIDADES...</div><div id="modalActividadSuggest" class="modal-suggest"></div><label class="form-label">LABOR</label><input id="modalLaborInput" name="subgrupo" class="form-control mb-1" placeholder="Seleccione labor según actividad" required autocomplete="off"><datalist id="modal_labor_list"></datalist><div id="modalLaborSuggest" class="modal-suggest"></div><label class="form-label">CONSUMIDOR (opcional)</label><input id="modalConsumidor" name="labor" class="form-control mb-1" placeholder="Consumidor / zona / campo"><datalist id="modal_consumidor_list"></datalist><div id="modalConsumidorSuggest" class="modal-suggest"></div><label class="form-label">RESPONSABLE</label><input name="responsable" class="form-control mb-2" placeholder="APELLIDOS Y NOMBRES" value="{{h.responsable}}"><div class="row g-2"><div class="col-6"><label class="form-label">TURNO</label><select name="turno" class="form-select"><option>DIA</option><option>NOCHE</option></select></div><div class="col-6"><label class="form-label">TIPO</label><select name="tipo_tareo" class="form-select"><option>JORNAL</option><option>RENDIMIENTO</option></select></div></div></div><div class="modal-footer"><button class="btn btn-green w-100" type="submit">CREAR LABOR</button></div></form></div></div></div>
    <div class="modal fade" id="modalCopiar" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('copiar_labor_hoja', hoja_id=h.id, tab=tab)}}"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-files"></i> Copiar labor existente</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small">Selecciona el documento/labor que deseas copiar. No se copiará nada hasta presionar <b>COPIAR SELECCIONADO</b>.</div><div class="copy-list">{% for l in labores %}<label class="d-block mb-2"><input type="radio" name="labor_id_origen" value="{{l.id}}" required> <b>{{l.labor}}</b><br><span class="small text-muted">{{l.grupo}} / {{l.subgrupo}} / {{l.turno}} / {{l.tipo_tareo}}</span></label>{% endfor %}</div><label class="form-label mt-2">Nuevo nombre de labor (opcional)</label><input name="labor_nueva" class="form-control" placeholder="Dejar vacío para copiar igual"></div><div class="modal-footer"><button class="btn btn-green w-100">COPIAR SELECCIONADO</button></div></form></div></div></div>
    <div class="modal fade" id="modalBuscar" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-search"></i> Buscar trabajador</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><input id="buscarDni" class="form-control mb-2" placeholder="DNI / QR / código barras"><button class="btn btn-green w-100" onclick="buscarTrabajadorLibre()">BUSCAR</button><div id="buscarResultado" class="alert alert-light border mt-2">Esperando búsqueda.</div></div></div></div></div>
    <div class="modal fade" id="modalHora" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('fijar_horario_hoja', hoja_id=h.id, tab=tab)}}"><input type="hidden" name="labor_id" value="{{selected_labor_id}}"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-clock"></i> Fijar horario obligatorio</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-warning small"><b>Obligatorio:</b> fija el horario antes de tarear trabajadores. Formato 24 horas (ej. 06:30, 16:30, 22:00, 06:00).</div><div class="touch-clock-panel"><div class="clock-24-hint">Toque un campo, mueva la manija o digite la hora en formato 24:00.</div><div id="touchClockValue" class="time-display">06:30</div><input id="timeSlider24" class="time-slider" type="range" min="0" max="1435" step="5" value="390"><div id="clockPickFields" class="touch-clock-picks"><button type="button" data-target="horaInicioDefault" class="active">Inicio trabajo</button><button type="button" data-target="horaFinDefault">Fin trabajo</button><button type="button" data-target="refInicioDefault">Inicio refrigerio</button><button type="button" data-target="refFinDefault">Fin refrigerio</button></div></div><div class="row g-2"><div class="col-6"><label class="form-label">Inicio trabajo</label><input name="hora_inicio_default" id="horaInicioDefault" type="text" class="form-control locked-input" value="{{h.hora_inicio_default or '06:30'}}" required></div><div class="col-6"><label class="form-label">Fin trabajo</label><input name="hora_fin_default" id="horaFinDefault" type="text" class="form-control locked-input" value="{{h.hora_fin_default or '16:30'}}" required></div><div class="col-6"><label class="form-label">Inicio refrigerio</label><input name="ref_inicio_default" id="refInicioDefault" type="text" class="form-control locked-input" value="{{h.ref_inicio_default or '12:00'}}" required></div><div class="col-6"><label class="form-label">Fin refrigerio</label><input name="ref_fin_default" id="refFinDefault" type="text" class="form-control locked-input" value="{{h.ref_fin_default or '13:00'}}" required></div></div><div id="horarioCoherencia" class="field-help mt-2">El refrigerio debe quedar dentro de la jornada.</div><button class="btn btn-green w-100 mt-3" type="submit">FIJAR HORARIO</button></div></form></div></div></div>
    <div class="modal fade" id="modalEditTareo" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form id="frmEditTareo" method="post"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-pencil-square"></i> Editar horas del trabajador</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small">Solo se puede editar si la hoja aún no fue enviada.</div><div class="row g-2"><div class="col-6"><label class="form-label">Hora inicio</label><input id="editHi" name="hora_inicio" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div><div class="col-6"><label class="form-label">Hora fin</label><input id="editHf" name="hora_fin" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div><div class="col-6"><label class="form-label">Ref. ini</label><input id="editRi" name="ref_inicio" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div><div class="col-6"><label class="form-label">Ref. fin</label><input id="editRf" name="ref_fin" class="form-control" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div></div></div><div class="modal-footer"><button class="btn btn-green w-100">GUARDAR CAMBIOS</button></div></form></div></div></div>

    <div class="modal fade" id="modalRegistro" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('guardar_registro_hoja', hoja_id=h.id, tab='trabajadores')}}" id="frmTrab"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-person-plus"></i> Registrar trabajador</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="scan-box mb-2"><label class="form-label">DNI / QR / CÓDIGO BARRAS</label><div class="input-group"><input name="dni" id="dniTrab" class="form-control" placeholder="Escanee o digite DNI" autocomplete="off" inputmode="numeric" maxlength="30" oninput="autoDetectarDniInline(this)" onkeyup="autoDetectarDniInline(this)" onchange="autoDetectarDniInline(this)"><button type="button" class="btn btn-green" onclick="abrirScanner('readerTrab','dniTrab')"><i class="bi bi-upc-scan"></i></button></div><div id="readerTrab" style="display:none;margin-top:8px"></div><div id="dniStatus" class="mt-2 field-help">Escanee o digite DNI: al completar 8 dígitos se agregará al pre-registro con sonido.</div><input type="hidden" name="dnis_masivos" id="dnisMasivos"><div class="queue-title">PRE-REGISTRO DE TRABAJADORES</div><div id="workerQueue" class="worker-queue"><div class="text-muted small text-center">Aún no hay trabajadores detectados.</div></div></div><label class="form-label">LABOR SELECCIONADA</label><input type="hidden" name="labor_id" value="{{selected_labor_id}}"><div class="form-control mb-2" style="height:auto;min-height:37px;background:#f8fff9;color:#166534;font-weight:900">{% if selected_labor %}{{selected_labor.grupo}} / {{selected_labor.subgrupo}} / {{selected_labor.labor}} / {{selected_labor.turno}} / {{selected_labor.tipo_tareo}}{% else %}PRIMERO SELECCIONA UNA LABOR{% endif %}</div><input name="turno" id="turnoTrab" type="hidden" value="DIA"><input name="tipo_tareo" type="hidden" value="JORNAL"><input name="hora_inicio" id="horaInicioTrab" type="hidden" value="{{h.hora_inicio_default or '06:30'}}"><input name="hora_fin" id="horaFinTrab" type="hidden" value="{{h.hora_fin_default or '16:30'}}"><input name="ref_inicio" id="refInicioTrab" type="hidden" value="{{h.ref_inicio_default or '12:00'}}"><input name="ref_fin" id="refFinTrab" type="hidden" value="{{h.ref_fin_default or '13:00'}}"><input name="horas" id="horasTrab" type="hidden" value="0"><input name="cantidad" type="hidden" value="0.00"><div id="horarioActivoTxt" class="alert {{'alert-success' if h.horario_fijado else 'alert-warning'}} small mt-2 mb-0"><b>Horario activo:</b> {{h.hora_inicio_default or 'NO FIJADO'}} - {{h.hora_fin_default or 'NO FIJADO'}} / Refrigerio {{h.ref_inicio_default or '--:--'}} - {{h.ref_fin_default or '--:--'}}. {% if not h.horario_fijado %}<b>Primero fija el horario desde el icono de reloj.</b>{% endif %}</div></div><div class="modal-footer"><button class="btn btn-green w-100">GUARDAR TRABAJADORES</button></div></form></div></div></div>
    <div class="modal fade" id="modalAvance" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form method="post" action="{{url_for('guardar_registro_hoja', hoja_id=h.id, tab='rendimiento')}}" id="frmAvance"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-upc-scan"></i> Registrar avance / lectura</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="scan-box mb-2"><label class="form-label">DNI / QR / CÓDIGO BARRAS DEL TRABAJADOR</label><div class="input-group"><input name="dni" id="dniAvance" class="form-control" placeholder="Escanee o digite DNI" required autocomplete="off" inputmode="numeric"><button type="button" class="btn btn-green" onclick="abrirScanner('readerAvance','dniAvance')"><i class="bi bi-upc-scan"></i></button></div><div id="avanceTrabStatus" class="field-help mt-2">Debe estar registrado en Trabajadores de esta labor.</div><div id="readerAvance" style="display:none;margin-top:8px"></div><label class="form-label mt-2">LECTOR AVANCE / QR / CÓDIGO BARRAS</label><div class="input-group"><input name="codigo_avance" id="codigoAvance" class="form-control" placeholder="OPCIONAL: CÓDIGO DE AVANCE / BALDE / ETIQUETA"><button type="button" class="btn btn-green" onclick="abrirScanner('readerCodigoAvance','codigoAvance')"><i class="bi bi-upc-scan"></i></button></div><div id="readerCodigoAvance" style="display:none;margin-top:8px"></div><div id="cantidadDetectada" class="scan-ok mt-2" style="display:none">Cantidad detectada: <b>1.00</b></div></div><label class="form-label">LABOR SELECCIONADA</label><input type="hidden" name="labor_id" value="{{selected_labor_id}}" id="avanceLaborId"><div class="form-control mb-2" style="height:auto;min-height:37px;background:#f8fff9;color:#166534;font-weight:900">{% if selected_labor %}{{selected_labor.labor}} / {{selected_labor.turno}} / {{selected_labor.tipo_tareo}}{% else %}PRIMERO SELECCIONA UNA LABOR{% endif %}</div><input type="hidden" name="avance_pre_json" id="avancePreJson"><div class="row g-2"><div class="col-12"><label class="form-label">CANTIDAD</label><input name="cantidad" id="cantidadAvance" type="number" step="0.01" class="form-control" value="1.00"></div><input name="a_noct" id="aNoctAvance" type="hidden" value="0.00"><div class="col-6"><label class="form-label">UNIDAD</label><select name="unidad" class="form-select"><option>BALDE</option><option>KG</option><option>JABA</option><option>UNIDAD</option></select></div><div class="col-6"><label class="form-label">MÉTODO</label><select name="metodo" class="form-select"><option>QR/CÓDIGO</option><option>DIGITACIÓN</option><option>LECTOR USB</option></select></div></div><div class="queue-title">PRE-REGISTRO DE AVANCES</div><div id="avanceQueue" class="worker-queue"><div class="text-muted small text-center">Aún no hay avances pre-registrados.</div></div></div><div class="modal-footer"><button class="btn btn-green w-100" id="btnGuardarAvance">GUARDAR AVANCE FINAL</button></div></form></div></div></div>
    <div class="modal fade" id="modalEditAvance" tabindex="-1"><div class="modal-dialog modal-dialog-centered"><div class="modal-content"><form id="frmEditAvance" method="post"><div class="modal-header"><h5 class="modal-title fw-bold text-success"><i class="bi bi-pencil-square"></i> Editar avance</h5><button class="btn-close" data-bs-dismiss="modal"></button></div><div class="modal-body"><div class="alert alert-light border small">Solo se puede editar si la hoja aún no fue enviada.</div><div class="row g-2"><div class="col-12"><label class="form-label">CANTIDAD</label><input id="editCantAvance" name="cantidad" type="number" step="0.01" class="form-control" required></div><input id="editNoctAvance" name="a_noct" type="hidden" value="0.00"></div></div><div class="modal-footer"><button class="btn btn-green w-100">GUARDAR CAMBIOS</button></div></form></div></div></div>

    <script>
(function(){
  'use strict';
  const MAESTROS_DET={{ maestros_json|safe }};
  const $=(id)=>document.getElementById(id);
  const pad=(n)=>String(Number(n)||0).padStart(2,'0');
  const sleep=(ms)=>new Promise(r=>setTimeout(r,ms));
  function playOk(){try{beep();}catch(e){}}
  function onlyDni(v){
    const raw=String(v||'');
    const m=raw.match(/(?:^|\D)(\d{8})(?:\D|$)/);
    const d=m?m[1]:raw.replace(/\D/g,'');
    return d.length>=8?d.slice(-8):d;
  }
  window.limpiarDni=onlyDni;

  // ================== BUSCADOR Y QR/CÓDIGO ==================
  window.buscarTrabajadorLibre=async function(){
    const inp=$('buscarDni'), box=$('buscarResultado');
    const dni=onlyDni(inp?inp.value:'');
    if(inp)inp.value=dni;
    if(!box)return;
    if(dni.length!==8){box.className='alert alert-warning mt-2';box.textContent='Ingrese DNI válido de 8 dígitos.';return;}
    box.className='alert alert-light border mt-2'; box.innerHTML='Buscando <b>'+dni+'</b>...';
    try{
      const r=await fetch('/api/trabajador/'+encodeURIComponent(dni),{cache:'no-store',credentials:'same-origin'});
      const j=await r.json();
      if(!j.ok){box.className='alert alert-danger mt-2'; box.textContent=j.msg||'DNI no encontrado.'; return;}
      const t=j.trabajador||{};
      box.className='alert alert-success mt-2';
      box.innerHTML='<b>'+(t.trabajador||'TRABAJADOR')+'</b><br>'+dni+' · '+(t.cargo||'')+' · '+(t.area||'');
      playOk();
    }catch(e){box.className='alert alert-danger mt-2';box.textContent='No se pudo consultar la base de trabajadores.';}
  };

  let scanner=null;
  function playBad(){try{playOk();}catch(e){}}
  function loadQrLib(){
    return new Promise(resolve=>{
      if(window.Html5Qrcode) return resolve(true);
      const old=document.querySelector('script[data-qr-loader="1"]');
      if(old){old.addEventListener('load',()=>resolve(true)); old.addEventListener('error',()=>resolve(false)); return;}
      const sc=document.createElement('script');
      sc.dataset.qrLoader='1';
      sc.src='https://cdn.jsdelivr.net/npm/html5-qrcode@2.3.8/html5-qrcode.min.js';
      sc.onload=()=>resolve(!!window.Html5Qrcode);
      sc.onerror=()=>resolve(false);
      document.head.appendChild(sc);
    });
  }
  window.cerrarScannerActivo=async function(){
    try{ if(scanner){ await scanner.stop().catch(()=>{}); await scanner.clear().catch(()=>{}); scanner=null; } }catch(e){}
    try{ if(window.__qr250){ await window.__qr250.stop().catch(()=>{}); await window.__qr250.clear().catch(()=>{}); window.__qr250=null; } }catch(e){}
    document.querySelectorAll('[id^="reader"]').forEach(el=>{el.style.display='none'; el.innerHTML='';});
  };
  function scannerCloseHtml(){
    return '<button type="button" class="scanner-close-x" title="Apagar cámara" onclick="cerrarScannerActivo()">×</button>';
  }
  window.abrirScanner=async function(readerId,inputId){
    const el=$(readerId), input=$(inputId); if(!el||!input)return;
    await window.cerrarScannerActivo();
    el.style.display='block';
    el.style.position='relative';
    el.innerHTML=scannerCloseHtml()+'<div class="p-2 text-success fw-bold">Abriendo cámara...</div>';
    if(location.protocol!=='https:' && location.hostname!=='localhost' && location.hostname!=='127.0.0.1'){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">La cámara requiere HTTPS. Abre la app desde Render con https://</div>';
      playBad(); return;
    }
    if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">Este navegador no permite cámara. Usa Chrome actualizado.</div>';
      playBad(); return;
    }
    const libOk=await loadQrLib();
    if(!libOk){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">No cargó la librería del lector. Revisa internet o CDN.</div>';
      playBad(); return;
    }
    try{
      let cfg={};
      if(window.Html5QrcodeSupportedFormats){
        cfg={formatsToSupport:[
          Html5QrcodeSupportedFormats.QR_CODE,
          Html5QrcodeSupportedFormats.CODE_128,
          Html5QrcodeSupportedFormats.CODE_39,
          Html5QrcodeSupportedFormats.EAN_13,
          Html5QrcodeSupportedFormats.EAN_8,
          Html5QrcodeSupportedFormats.UPC_A,
          Html5QrcodeSupportedFormats.UPC_E
        ]};
      }
      scanner=new Html5Qrcode(readerId,cfg);
      await scanner.start({facingMode:{ideal:'environment'}},{fps:12,qrbox:{width:240,height:160},aspectRatio:1.333},async decoded=>{
        let val=String(decoded||'').trim();
        if(inputId==='dniTrab'||inputId==='dniAvance') val=onlyDni(val);
        input.value=val;
        input.dispatchEvent(new Event('input',{bubbles:true}));
        input.dispatchEvent(new Event('change',{bubbles:true}));
        if(inputId==='dniTrab') await procesarDni(val,true);
        else playOk();
        await window.cerrarScannerActivo();
      },()=>{});
      if(!el.querySelector('.scanner-close-x')) el.insertAdjacentHTML('afterbegin', scannerCloseHtml());
    }catch(err){
      el.innerHTML=scannerCloseHtml()+'<div class="scan-bad mt-2">No se pudo activar cámara. Permite cámara en el candado del navegador y vuelve a intentar.</div>';
      playBad();
    }
  };

  // ================== DNI AUTOMÁTICO REAL ==================
  const workerMap=new Map();
  window.workerMap=workerMap;
  window.renderQueue=function(){
    const q=$('workerQueue'), h=$('dnisMasivos'); if(!q||!h)return;
    h.value=[...workerMap.keys()].join(',');
    if(workerMap.size===0){q.innerHTML='<div class="text-muted small text-center">Aún no hay trabajadores detectados.</div>';return;}
    q.innerHTML=[...workerMap.entries()].map(([dni,n])=>
      '<div class="queue-item"><div><b>'+dni+'</b><br><span>'+n+'</span></div><button type="button" class="btn btn-sm btn-outline-danger" onclick="workerMap.delete(\''+dni+'\');renderQueue();">×</button></div>'
    ).join('');
  };
  function dniStatus(kind,html){
    const st=$('dniStatus'); if(!st)return;
    st.className=(kind==='ok'?'scan-ok mt-2 flash':kind==='bad'?'scan-bad mt-2 flash':'mt-2 field-help');
    st.innerHTML=html;
  }
  let dniTimer=null, dniBusy=false, dniLast='';
  async function procesarDni(valor, forzar=false){
    const inp=$('dniTrab'); if(!inp)return;
    const dni=onlyDni(valor || inp.value);
    if(dni.length<8){ if(forzar) dniStatus('help','Escanee o digite DNI: al completar 8 dígitos se agregará al pre-registro con sonido.'); return; }
    inp.value=dni;
    if(!forzar && dni===dniLast)return;
    if(dniBusy){ clearTimeout(dniTimer); dniTimer=setTimeout(()=>procesarDni(dni,true),120); return; }
    dniBusy=true; dniLast=dni;
    dniStatus('ok','Buscando DNI <b>'+dni+'</b> en base trabajadores...');
    try{
      const r=await fetch('/api/trabajador/'+encodeURIComponent(dni),{cache:'no-store',credentials:'same-origin'});
      let j={ok:false,msg:'Respuesta inválida'}; try{j=await r.json();}catch(e){}
      if(!j.ok){
        dniStatus('bad','✕ '+(j.msg||'DNI no encontrado en base trabajadores')+' <b>'+dni+'</b>');
        playOk(); inp.select(); return;
      }
      const t=j.trabajador||{}; const nombre=t.trabajador||t.nombres||t.nombre||'TRABAJADOR';
      if(!workerMap.has(dni)){workerMap.set(dni,nombre); window.renderQueue();}
      dniStatus('ok','✓ Reconocido automáticamente: <b>'+nombre+'</b> · '+dni);
      playOk();
      await sleep(180); inp.value=''; dniLast=''; inp.focus();
    }catch(e){
      dniStatus('bad','Error consultando trabajador. Revisa conexión o sesión.');
    }finally{dniBusy=false;}
  }
  window.autoDetectarDniInline=function(el){
    const inp=el||$('dniTrab'); if(!inp)return;
    clearTimeout(dniTimer);
    dniTimer=setTimeout(()=>procesarDni(inp.value,false),40);
  };
  function instalarDniAuto(){
    const inp=$('dniTrab'); if(!inp || inp.dataset.finalAuto==='1')return; inp.dataset.finalAuto='1';
    ['input','keyup','change','paste','blur'].forEach(ev=>inp.addEventListener(ev,()=>setTimeout(()=>procesarDni(inp.value,ev!=='input'), ev==='paste'?80:5), true));
    inp.addEventListener('keydown',e=>{ if(e.key==='Enter'||e.key==='Tab'){procesarDni(inp.value,true); if(e.key==='Enter')e.preventDefault();} }, true);
    setInterval(()=>{const x=$('dniTrab'); if(x && x.value && onlyDni(x.value).length>=8) procesarDni(x.value,false);},250);
  }

  // ================== HORARIO TÁCTIL / CURSOR PC POR DESLIZADOR 24H ==================
  const IDS=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  let campoActivo='horaInicioDefault';
  function toMin(v){const p=String(v||'00:00').split(':');let h=parseInt(p[0]||0,10),m=parseInt(p[1]||0,10); if(isNaN(h))h=0;if(isNaN(m))m=0;return Math.max(0,Math.min(1435,h*60+m));}
  function minToTime(m){m=((parseInt(m||0,10)%1440)+1440)%1440;return pad(Math.floor(m/60))+':'+pad(m%60);}
  function horasNetas(hi,hf,ri,rf){let a=toMin(hi), b=toMin(hf); if(b<=a)b+=1440; let total=b-a; if(ri&&rf){let c=toMin(ri), d=toMin(rf); if(d<=c)d+=1440; if(b>1440&&c<a){c+=1440;d+=1440;} total-=Math.max(0,Math.min(b,d)-Math.max(a,c));} return (Math.max(0,total)/60).toFixed(2);}
  function activeInput(){return $(campoActivo)||$('horaInicioDefault');}
  function sincHorario(){const hi=$('horaInicioDefault')?.value||'06:30', hf=$('horaFinDefault')?.value||'16:30', ri=$('refInicioDefault')?.value||'12:00', rf=$('refFinDefault')?.value||'13:00'; [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,v])=>{const e=$(id); if(e)e.value=v;}); const h=horasNetas(hi,hf,ri,rf); const ht=$('horasTrab'); if(ht)ht.value=h; const box=$('horarioActivoTxt'); if(box)box.innerHTML='<b>Horario activo:</b> '+hi+' - '+hf+' / Refrigerio '+ri+' - '+rf+' / H.Normal '+h+'.';}
  function pintarReloj(){const input=activeInput(), tv=$('touchClockValue'), sl=$('timeSlider24'), pills=$('clockPickFields'); if(tv&&input)tv.textContent=input.value; if(sl&&input)sl.value=toMin(input.value); if(pills)pills.querySelectorAll('button').forEach(b=>b.classList.toggle('active',b.dataset.target===campoActivo));}
  function setCampo(id){campoActivo=id;pintarReloj();const el=$(id);if(el){try{el.focus({preventScroll:true});}catch(_){}}}
  function setDesdeMinutos(m){const i=activeInput(); if(!i)return; i.value=minToTime(Math.round(Number(m||0)/5)*5); sincHorario(); pintarReloj();}
  function instalarReloj(){
    const sl=$('timeSlider24'), pills=$('clockPickFields');
    if(pills){pills.querySelectorAll('button').forEach(b=>{b.onclick=(ev)=>{ev.preventDefault(); setCampo(b.dataset.target);}; b.onpointerdown=(ev)=>{ev.preventDefault(); setCampo(b.dataset.target);};});}
    IDS.forEach(id=>{const e=$(id); if(e){e.readOnly=true; e.tabIndex=0; e.onclick=()=>setCampo(id); e.onpointerdown=()=>setCampo(id); e.onfocus=()=>setCampo(id);}});
    if(sl){
      sl.style.pointerEvents='auto'; sl.style.touchAction='pan-x';
      const handler=(ev)=>{setDesdeMinutos(sl.value);};
      ['input','change','mousemove','touchmove'].forEach(ev=>{sl.addEventListener(ev,handler,{passive:true});});
      sl.onpointerdown=(ev)=>{setDesdeMinutos(sl.value);};
      sl.onpointerup=(ev)=>{setDesdeMinutos(sl.value); playOk();};
      sl.onclick=(ev)=>{setDesdeMinutos(sl.value);};
    }
    sincHorario(); pintarReloj();
  }
  window.aplicarHorarioRegistro=function(){sincHorario();pintarReloj();playOk();};
  window.tocarHora=function(delta){setDesdeMinutos(toMin(activeInput()?.value)+delta*60);};
  window.tocarMin=function(delta){setDesdeMinutos(toMin(activeInput()?.value)+delta);};
  window.setCampoHorario=setCampo;
  
  window.abrirEditarTareo=function(id,hi,hf,ri,rf){
    const f=$('frmEditTareo'); if(!f){alert('No se pudo abrir editor de horario.');return;}
    document.querySelectorAll('.card-menu.show').forEach(m=>m.classList.remove('show'));
    f.action='/tareo/'+id+'/editar-horas';
    $('editHi').value=hi||'06:30'; $('editHf').value=hf||'16:30'; $('editRi').value=ri||'12:00'; $('editRf').value=rf||'13:00'; setTimeout(()=>{try{window.instalarSelectorHorarioIOS&&window.instalarSelectorHorarioIOS();}catch(e){}},120);
    const modalEl=$('modalEditTareo');
    if(!modalEl){ window.location.href='/tareo/'+id+'/editar-horas-form'; return; }
    modalEl.style.zIndex='20000';
    if(window.bootstrap&&bootstrap.Modal){bootstrap.Modal.getOrCreateInstance(modalEl).show();}
    else {modalEl.style.display='block'; modalEl.classList.add('show'); modalEl.removeAttribute('aria-hidden');}
  };

  document.addEventListener('click',e=>{
    const b=e.target.closest && e.target.closest('.btn-edit-tareo');
    if(!b)return;
    e.preventDefault(); e.stopPropagation();
    window.abrirEditarTareo(b.dataset.id,b.dataset.hi,b.dataset.hf,b.dataset.ri,b.dataset.rf);
  },true);


  // PATCH 252: valida coherencia antes de guardar horario general o editar horario del trabajador
  function normHHMM(v){
    v=String(v||'').trim();
    if(/^\d{1,2}:\d{2}$/.test(v)){let [h,m]=v.split(':'); return String(Number(h)).padStart(2,'0')+':'+m;}
    const d=v.replace(/\D/g,'');
    if(d.length===3) return '0'+d[0]+':'+d.slice(1);
    if(d.length>=4) return d.slice(0,2)+':'+d.slice(2,4);
    return v;
  }
  function minHH(v){v=normHHMM(v); const p=v.split(':'); return (Number(p[0]||0)*60)+Number(p[1]||0);}
  function horarioOkClient(hi,hf,ri,rf){
    hi=normHHMM(hi); hf=normHHMM(hf); ri=normHHMM(ri); rf=normHHMM(rf);
    const re=/^([01][0-9]|2[0-3]):[0-5][0-9]$/;
    if(![hi,hf,ri,rf].every(x=>re.test(x))) return {ok:false,msg:'Usa formato 24 horas HH:MM. Ejemplo: 06:30, 16:30, 22:00.'};
    let a=minHH(hi), b=minHH(hf), c=minHH(ri), d=minHH(rf);
    if(b<=a) b+=1440;
    if(d<=c) d+=1440;
    if(b>1440 && c<a){c+=1440; d+=1440;}
    if(!(a<=c && c<d && d<=b)) return {ok:false,msg:'El refrigerio debe quedar dentro del horario de inicio y fin de trabajo.'};
    if((d-c)>(b-a)) return {ok:false,msg:'El refrigerio no puede ser mayor que la jornada.'};
    // Permite jornadas en formato 24 horas e incluso cruce de medianoche.
    // La única regla bloqueante es que el refrigerio quede dentro del inicio y fin de trabajo.
    return {ok:true,msg:'Horario coherente.'};
  }
  function validarFormHorario(form, ids){
    const vals=ids.map(id=>$(id)); if(vals.some(x=>!x)) return true;
    vals.forEach(x=>x.value=normHHMM(x.value));
    const r=horarioOkClient(vals[0].value, vals[1].value, vals[2].value, vals[3].value);
    const box=$('horarioCoherencia');
    if(box){box.className=r.ok?'scan-ok mt-2':'scan-bad mt-2'; box.innerHTML=r.msg;}
    if(!r.ok){alert(r.msg); return false;}
    return true;
  }
  document.addEventListener('submit',e=>{
    const f=e.target;
    if(f && f.action && f.action.includes('/fijar-horario/')){
      if(!validarFormHorario(f,['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'])){e.preventDefault(); e.stopPropagation(); return false;}
    }
    if(f && f.id==='frmEditTareo'){
      const vals=['editHi','editHf','editRi','editRf'].map(id=>$(id)); vals.forEach(x=>{if(x)x.value=normHHMM(x.value);});
      const r=horarioOkClient(vals[0]?.value, vals[1]?.value, vals[2]?.value, vals[3]?.value);
      if(!r.ok){e.preventDefault(); e.stopPropagation(); alert(r.msg); return false;}
    }
  },true);


  // ================== MAYÚSCULAS Y MAESTROS ACTIVIDAD/LABOR/CONSUMIDOR ==================
  document.addEventListener('input',e=>{const el=e.target;if(el && (el.tagName==='INPUT'||el.tagName==='TEXTAREA') && el.type!=='password' && el.type!=='date' && el.type!=='number' && el.id!=='dniTrab' && el.id!=='dniAvance' && el.id!=='buscarDni'){const p=el.selectionStart; el.value=String(el.value||'').toUpperCase(); try{el.setSelectionRange(p,p)}catch(_){}}},true);
  let MAESTROS_CACHE = (typeof MAESTROS_DET !== 'undefined' && Array.isArray(MAESTROS_DET)) ? MAESTROS_DET.slice() : [];
  const DEMO_MAESTROS=[{desc_actividad:'ADMINISTRACION',desc_labor:'LABOR ADMINISTRATIVA',desc_consumidor:'OFICINA CENTRAL'},{desc_actividad:'ADMISION',desc_labor:'CONTROL DOCUMENTARIO',desc_consumidor:'OFICINA 01'},{desc_actividad:'COSECHA',desc_labor:'COSECHA MANUAL',desc_consumidor:'CAMPO 01'},{desc_actividad:'COSECHA',desc_labor:'COSECHA SELECTIVA',desc_consumidor:'CAMPO 02'},{desc_actividad:'PODA',desc_labor:'PODA SANITARIA',desc_consumidor:'LOTE 01'}];
  const uniq=(a)=>[...new Set((a||[]).map(x=>String(x||'').trim()).filter(Boolean))].sort((a,b)=>a.localeCompare(b));
  function norm(v){return String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toUpperCase().trim();}
  function firstVal(x,names){for(const n of names){if(x && x[n]!==undefined && x[n]!==null && String(x[n]).trim()!=='')return String(x[n]).trim().toUpperCase();}return '';}
  function normalizaFila(x){return {desc_actividad:firstVal(x,['desc_actividad','actividad','grupo','ACTIVIDAD','DESCRIPCION_ACTIVIDAD','DESCRIPCION ACTIVIDAD','DESC ACTIVIDAD','cod_actividad','COD_ACTIVIDAD']),cod_actividad:firstVal(x,['cod_actividad','codigo_actividad','COD ACTIVIDAD','COD_ACTIVIDAD']),desc_labor:firstVal(x,['desc_labor','labor','subgrupo','LABOR','DESCRIPCION_LABOR','DESCRIPCION LABOR','DESC LABOR','cod_labor','COD_LABOR']),cod_labor:firstVal(x,['cod_labor','codigo_labor','COD LABOR','COD_LABOR']),desc_consumidor:firstVal(x,['desc_consumidor','consumidor','CONSUMIDOR','zona','campo','ZONA','CAMPO','DESCRIPCION_CONSUMIDOR','DESCRIPCION CONSUMIDOR','cod_consumidor','COD_CONSUMIDOR']),cod_consumidor:firstVal(x,['cod_consumidor','codigo_consumidor','COD CONSUMIDOR','COD_CONSUMIDOR'])};}
  function setMasterStatus(msg,bad=false){const st=$('modalMasterStatus'); if(st){st.className='master-status'+(bad?' bad':'');st.textContent=msg;}}
  async function cargarMaestrosSiHaceFalta(){
    try{
      const r=await fetch('/api/actividades-maestras?ts='+Date.now(),{cache:'no-store',credentials:'same-origin'});
      const j=await r.json();
      if(j && j.ok && Array.isArray(j.data)){MAESTROS_CACHE=j.data.map(normalizaFila).filter(x=>x.desc_actividad||x.desc_labor||x.desc_consumidor);}
    }catch(e){setMasterStatus('NO SE PUDO CONECTAR A LA API DE ACTIVIDADES',true);}
    if(!MAESTROS_CACHE.length){MAESTROS_CACHE=DEMO_MAESTROS.slice(); setMasterStatus('SIN DATA REAL EN API: USANDO DEMO TEMPORAL',true);} else {setMasterStatus('ACTIVIDADES CARGADAS: '+MAESTROS_CACHE.length+' REGISTROS');}
  }
  function fillDL(id,arr){const dl=$(id); if(!dl)return; dl.innerHTML=''; arr.slice(0,500).forEach(v=>{const o=document.createElement('option');o.value=v;dl.appendChild(o);});}
  function showSuggest(id,input,arr,cb){const box=$(id); if(!box||!input)return; const q=norm(input.value); const vals=(q?arr.filter(v=>norm(v).includes(q)):arr).slice(0,20); if(!vals.length){box.style.display='none';box.innerHTML='';return;} box.innerHTML=vals.map(v=>'<div>'+String(v).replace(/</g,'&lt;')+'</div>').join(''); box.style.display='block'; [...box.children].forEach(div=>{div.onmousedown=(ev)=>ev.preventDefault(); div.onclick=()=>{input.value=div.textContent.toUpperCase();box.style.display='none';cb&&cb();};});}
  async function instalarMaestros(){
    await cargarMaestrosSiHaceFalta();
    const a=$('modalActividad'), l=$('modalLaborInput'), c=$('modalConsumidor'); if(!a||!l)return;
    const acts=uniq(MAESTROS_CACHE.map(x=>x.desc_actividad||x.cod_actividad)); fillDL('modal_actividad_list',acts);
    const rowsA=()=>{const q=norm(a.value);return MAESTROS_CACHE.filter(x=>!q||norm(x.desc_actividad).includes(q)||norm(x.cod_actividad).includes(q));};
    const laborVals=()=>uniq(rowsA().map(x=>x.desc_labor||x.cod_labor));
    const rowsL=()=>{const q=norm(l.value);return rowsA().filter(x=>!q||norm(x.desc_labor).includes(q)||norm(x.cod_labor).includes(q));};
    const consVals=()=>uniq(rowsL().map(x=>x.desc_consumidor||x.cod_consumidor));
    const refreshC=(show=false)=>{const vals=consVals(); fillDL('modal_consumidor_list',vals); if(c&&show)showSuggest('modalConsumidorSuggest',c,vals);};
    const refreshL=(show=false)=>{const vals=laborVals(); fillDL('modal_labor_list',vals); if(show)showSuggest('modalLaborSuggest',l,vals,()=>{refreshC(true); c&&c.focus();}); refreshC(false);};
    a.oninput=()=>{showSuggest('modalActividadSuggest',a,acts,()=>{l.value=''; if(c)c.value=''; refreshL(true);l.focus();});refreshL(true);};
    a.onfocus=()=>showSuggest('modalActividadSuggest',a,acts,()=>{l.value=''; refreshL(true);});
    l.oninput=()=>{showSuggest('modalLaborSuggest',l,laborVals(),()=>{if(c)c.value=''; refreshC(true);c&&c.focus();});refreshC(false);};
    l.onfocus=()=>showSuggest('modalLaborSuggest',l,laborVals(),()=>refreshC(true));
    if(c){c.oninput=()=>refreshC(true); c.onfocus=()=>refreshC(true);}
    document.addEventListener('click',ev=>{if(!ev.target.closest('#modalLabor'))['modalActividadSuggest','modalLaborSuggest','modalConsumidorSuggest'].forEach(id=>{const b=$(id);if(b)b.style.display='none';});},true);
    refreshL(false);
  }

  document.addEventListener('shown.bs.modal',e=>{
    if(e.target&&e.target.id==='modalRegistro'){instalarDniAuto();sincHorario();const i=$('dniTrab');if(i){setTimeout(()=>i.focus(),80);}}
    if(e.target&&e.target.id==='modalHora'){setTimeout(instalarReloj,60);}
    if(e.target&&e.target.id==='modalLabor'){instalarMaestros();}
  });
  document.addEventListener('DOMContentLoaded',()=>{instalarDniAuto();instalarReloj();instalarMaestros();});
  document.addEventListener('submit',e=>{if(e.target&&e.target.id==='frmTrab'){sincHorario(); window.renderQueue();}});
})();
</script>

<script>
/* === PARCHE REAL 247: maestros + horario deslizable robusto === */
(function(){
  'use strict';
  const $ = (id)=>document.getElementById(id);
  const norm = (v)=>String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toUpperCase().trim();
  const pad = (n)=>String(Number(n)||0).padStart(2,'0');
  function minToTime(m){m=Math.max(0,Math.min(1435,parseInt(m||0,10)));return pad(Math.floor(m/60))+':'+pad(m%60);}
  function toMin(v){let p=String(v||'00:00').split(':'),h=parseInt(p[0]||0,10),m=parseInt(p[1]||0,10); if(isNaN(h))h=0;if(isNaN(m))m=0;return Math.max(0,Math.min(1435,h*60+m));}
  function safeText(v){return String(v||'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));}

  // ---------- ACTIVIDAD / LABOR / CONSUMIDOR DESDE API ----------
  let maestros=[];
  const demo=[
    {desc_actividad:'ADMINISTRACION',desc_labor:'LABOR ADMINISTRATIVA',desc_consumidor:'OFICINA CENTRAL'},
    {desc_actividad:'ADMISION',desc_labor:'CONTROL DOCUMENTARIO',desc_consumidor:'OFICINA 01'},
    {desc_actividad:'COSECHA',desc_labor:'COSECHA MANUAL',desc_consumidor:'CAMPO 01'},
    {desc_actividad:'COSECHA',desc_labor:'COSECHA SELECTIVA',desc_consumidor:'CAMPO 02'},
    {desc_actividad:'PODA',desc_labor:'PODA SANITARIA',desc_consumidor:'LOTE 01'}
  ];
  function fila(x){
    const pick=(names)=>{for(const n of names){if(x&&x[n]!=null&&String(x[n]).trim())return norm(x[n]);}return '';};
    return {
      desc_actividad:pick(['desc_actividad','actividad','grupo','cod_actividad','ACTIVIDAD','DESCRIPCION_ACTIVIDAD','DESCRIPCION ACTIVIDAD']),
      desc_labor:pick(['desc_labor','labor','subgrupo','cod_labor','LABOR','DESCRIPCION_LABOR','DESCRIPCION LABOR']),
      desc_consumidor:pick(['desc_consumidor','consumidor','zona','campo','cod_consumidor','CONSUMIDOR','ZONA','CAMPO'])
    };
  }
  function unique(a){return [...new Set((a||[]).filter(Boolean))].sort((x,y)=>x.localeCompare(y));}
  function status(msg,bad){let s=$('modalMasterStatus'); if(s){s.className='master-status'+(bad?' bad':''); s.textContent=msg;}}
  function fillDatalist(id, arr){let dl=$(id); if(!dl)return; dl.innerHTML=''; arr.slice(0,500).forEach(v=>{let o=document.createElement('option');o.value=v;dl.appendChild(o);});}
  function showBox(id,input,vals,onpick){
    let b=$(id); if(!b||!input)return; let q=norm(input.value); let list=(q?vals.filter(v=>norm(v).includes(q)):vals).slice(0,30);
    if(!list.length){b.style.display='none'; b.innerHTML=''; return;}
    b.innerHTML=list.map(v=>'<div>'+safeText(v)+'</div>').join(''); b.style.display='block';
    [...b.children].forEach(d=>{d.onmousedown=e=>e.preventDefault(); d.onclick=()=>{input.value=d.textContent; b.style.display='none'; if(onpick)onpick();};});
  }
  async function cargarMaestros(){
    try{
      let r=await fetch('/api/actividades-maestras?ts='+Date.now(),{cache:'no-store',credentials:'same-origin'});
      let j=await r.json();
      maestros=(j&&j.ok&&Array.isArray(j.data)?j.data:[]).map(fila).filter(x=>x.desc_actividad||x.desc_labor||x.desc_consumidor);
    }catch(e){maestros=[];}
    if(!maestros.length){maestros=demo.map(fila); status('SIN DATA REAL EN API: USANDO DEMO TEMPORAL',true);} else status('ACTIVIDADES CARGADAS: '+maestros.length+' REGISTROS',false);
  }
  async function instalarMaestrosReal(){
    let a=$('modalActividad'), l=$('modalLaborInput'), c=$('modalConsumidor'); if(!a||!l)return;
    await cargarMaestros();
    const getActs=()=>unique(maestros.map(x=>x.desc_actividad));
    const rowsA=()=>{let q=norm(a.value);return maestros.filter(x=>!q||norm(x.desc_actividad).includes(q)||q.includes(norm(x.desc_actividad)));};
    const getLabs=()=>unique(rowsA().map(x=>x.desc_labor));
    const rowsL=()=>{let q=norm(l.value);return rowsA().filter(x=>!q||norm(x.desc_labor).includes(q)||q.includes(norm(x.desc_labor)));};
    const getCons=()=>unique(rowsL().map(x=>x.desc_consumidor));
    function refreshActividad(show=true){let vals=getActs();fillDatalist('modal_actividad_list',vals); if(show)showBox('modalActividadSuggest',a,vals,()=>{l.value=''; if(c)c.value=''; refreshLabor(true); setTimeout(()=>l.focus(),20);});}
    function refreshLabor(show=true){let vals=getLabs();fillDatalist('modal_labor_list',vals); if(show)showBox('modalLaborSuggest',l,vals,()=>{if(c)c.value=''; refreshConsumidor(true); setTimeout(()=>c&&c.focus(),20);}); refreshConsumidor(false);}
    function refreshConsumidor(show=true){let vals=getCons();fillDatalist('modal_consumidor_list',vals); if(c&&show)showBox('modalConsumidorSuggest',c,vals,()=>{});}
    a.oninput=()=>{a.value=norm(a.value); refreshActividad(true); refreshLabor(true);};
    a.onfocus=()=>refreshActividad(true);
    l.oninput=()=>{l.value=norm(l.value); refreshLabor(true);};
    l.onfocus=()=>refreshLabor(true);
    if(c){c.oninput=()=>{c.value=norm(c.value); refreshConsumidor(true);}; c.onfocus=()=>refreshConsumidor(true);}
    refreshActividad(false); refreshLabor(false);
  }

  // ---------- HORARIO POR DESLIZADOR: PC + TÁCTIL ----------
  const ids=['horaInicioDefault','horaFinDefault','refInicioDefault','refFinDefault'];
  let active='horaInicioDefault';
  function inputActive(){return $(active)||$('horaInicioDefault');}
  function paint(){let inp=inputActive(), sl=$('timeSlider24'), val=$('touchClockValue'); if(inp&&sl)sl.value=toMin(inp.value); if(inp&&val)val.textContent=inp.value; let box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>b.classList.toggle('active',b.dataset.target===active));}
  function setActive(id){active=id; paint(); let e=$(id); if(e)try{e.focus({preventScroll:true});}catch(_){} }
  function applySlider(){let inp=inputActive(); if(!inp)return; let sl=$('timeSlider24'); inp.value=minToTime(sl?sl.value:0); paint(); syncHidden();}
  function syncHidden(){
    const hi=$('horaInicioDefault')?.value||'06:30', hf=$('horaFinDefault')?.value||'16:30', ri=$('refInicioDefault')?.value||'12:00', rf=$('refFinDefault')?.value||'13:00';
    [['horaInicioTrab',hi],['horaFinTrab',hf],['refInicioTrab',ri],['refFinTrab',rf]].forEach(([id,v])=>{let e=$(id);if(e)e.value=v;});
    let a=toMin(hi),b=toMin(hf); if(b<=a)b+=1440; let c=toMin(ri),d=toMin(rf); if(d<=c)d+=1440; if(b>1440&&c<a){c+=1440;d+=1440;} let total=Math.max(0,(b-a)-Math.max(0,Math.min(b,d)-Math.max(a,c)))/60;
    let h=$('horasTrab'); if(h)h.value=total.toFixed(2);
    let txt=$('horarioActivoTxt'); if(txt)txt.innerHTML='<b>Horario activo:</b> '+hi+' - '+hf+' / Refrigerio '+ri+' - '+rf+' / H.Normal '+total.toFixed(2)+'.';
  }
  function instalarRelojReal(){
    let sl=$('timeSlider24');
    let box=$('clockPickFields'); if(box)[...box.querySelectorAll('button')].forEach(b=>{b.onclick=e=>{e.preventDefault();setActive(b.dataset.target);}; b.onpointerdown=e=>{setActive(b.dataset.target);};});
    ids.forEach(id=>{let e=$(id); if(e){e.readOnly=true; e.style.cursor='pointer'; e.onclick=()=>setActive(id); e.onpointerdown=()=>setActive(id); e.onfocus=()=>setActive(id);}});
    if(sl && sl.dataset.boundReal!=='1'){
      sl.dataset.boundReal='1'; sl.style.pointerEvents='auto'; sl.style.touchAction='none';
      ['input','change','pointermove','mousemove','touchmove','click','pointerup','touchend'].forEach(ev=>sl.addEventListener(ev,(e)=>{applySlider();},{passive:false}));
    }
    paint(); syncHidden();
  }
  window.setCampoHorario=setActive; window.aplicarHorarioRegistro=syncHidden;
  document.addEventListener('shown.bs.modal',e=>{if(e.target&&e.target.id==='modalLabor')setTimeout(instalarMaestrosReal,50); if(e.target&&e.target.id==='modalHora')setTimeout(instalarRelojReal,50);});
  document.addEventListener('DOMContentLoaded',()=>{setTimeout(()=>{instalarMaestrosReal(); instalarRelojReal();},100);});
})();
</script>

    """
    return render_page(body, h=h, tab=tab, tareos=tareos, lecturas=lecturas, labores=labores, registros=registros, horas_total=horas_total, rend_total=rend_total, maestros_json=js_master_options(get_actividades_maestras()), selected_labor_id=selected_labor_id, selected_labor=selected_labor)


@app.route('/hoja/<int:hoja_id>/copiar-labor/<tab>', methods=['POST'])
@login_required
def copiar_labor_hoja(hoja_id, tab):
    origen_id = request.form.get('labor_id_origen')
    lab = row_to_dict(execute('SELECT * FROM hoja_labores WHERE id=? AND hoja_id=?', (origen_id, hoja_id), fetchone=True))
    if not lab:
        flash('Selecciona una labor válida para copiar.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))
    labor_nueva = limpiar_texto(request.form.get('labor_nueva') or lab.get('labor'))
    execute('INSERT INTO hoja_labores(hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)',
            (hoja_id, lab.get('grupo'), lab.get('subgrupo'), labor_nueva, lab.get('turno'), lab.get('tipo_tareo'), lab.get('responsable'), now_str(), session.get('usuario')), commit=True)
    flash('Labor copiada correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))

@app.route('/hoja/<int:hoja_id>/labor/<tab>', methods=['POST'])
@login_required
def guardar_labor_hoja(hoja_id, tab):
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h: flash('Hoja no encontrada.', 'danger'); return redirect(url_for('hojas_tareo'))
    grupo = limpiar_texto(request.form.get('grupo') or h.get('grupo'))
    subgrupo = limpiar_texto(request.form.get('subgrupo') or h.get('subgrupo'))
    labor = limpiar_texto(request.form.get('labor') or h.get('labor'))
    responsable = limpiar_texto(request.form.get('responsable') or h.get('responsable'))
    turno = limpiar_texto(request.form.get('turno') or 'DIA')
    tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or 'JORNAL')
    if not grupo or not subgrupo:
        flash('Debe seleccionar Actividad y Labor.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))
    execute('INSERT INTO hoja_labores(hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,creado_en,creado_por) VALUES(?,?,?,?,?,?,?,?,?)',
            (hoja_id,grupo,subgrupo,labor,turno,tipo_tareo,responsable,now_str(),session.get('usuario')), commit=True)
    flash('Nueva labor creada dentro de la hoja.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))


@app.route('/hoja/<int:hoja_id>/fijar-horario/<tab>', methods=['POST'])
@login_required
def fijar_horario_hoja(hoja_id, tab):
    labor_id = request.form.get('labor_id') or request.args.get('labor_id') or ''
    def okhora(v):
        return bool(re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', str(v or '').strip()))
    hi = (request.form.get('hora_inicio_default') or '').strip().zfill(5)
    hf = (request.form.get('hora_fin_default') or '').strip().zfill(5)
    ri = (request.form.get('ref_inicio_default') or '').strip().zfill(5)
    rf = (request.form.get('ref_fin_default') or '').strip().zfill(5)
    if not all(okhora(x) for x in [hi,hf,ri,rf]):
        flash('Horario inválido. Usa formato 24 horas: HH:MM, por ejemplo 06:30 o 22:00.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id))
    coherente, msg_coh = horario_coherente(hi, hf, ri, rf)
    if not coherente:
        flash(msg_coh, 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id))
    execute('UPDATE hojas_tareo SET horario_fijado=1, hora_inicio_default=?, hora_fin_default=?, ref_inicio_default=?, ref_fin_default=? WHERE id=?',
            (hi,hf,ri,rf,hoja_id), commit=True)
    flash('Horario fijado correctamente para esta hoja de tareo.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id))

@app.route('/hoja/<int:hoja_id>/registro/<tab>', methods=['POST'])
@login_required
def guardar_registro_hoja(hoja_id, tab):
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))

    avance_pre_items = []
    if tab == 'rendimiento' and request.form.get('avance_pre_json'):
        try:
            raw_items = json.loads(request.form.get('avance_pre_json') or '[]')
            if isinstance(raw_items, list):
                for it in raw_items:
                    d = limpiar_dni((it or {}).get('dni'))
                    try:
                        cant_it = float((it or {}).get('cantidad') or 0)
                    except Exception:
                        cant_it = 0
                    metodo_it = limpiar_texto((it or {}).get('metodo') or request.form.get('metodo') or 'QR/CÓDIGO')
                    if len(d) == 8 and cant_it > 0:
                        avance_pre_items.append({'dni': d, 'cantidad': cant_it, 'a_noct': 0.0, 'metodo': metodo_it})
        except Exception:
            avance_pre_items = []

    dnis_raw = request.form.get('dnis_masivos') or request.form.get('dni') or ''
    dnis = []
    for part in re.split(r'[,;\s]+', dnis_raw):
        d = limpiar_dni(part)
        if len(d) == 8 and d not in dnis:
            dnis.append(d)
    if avance_pre_items:
        dnis = []
        for it in avance_pre_items:
            if it['dni'] not in dnis:
                dnis.append(it['dni'])
    if not dnis:
        flash('Debe digitar o escanear al menos un DNI válido.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab))
    if tab == 'trabajadores' and not int(h.get('horario_fijado') or 0):
        flash('Antes de tarear trabajadores debes fijar el horario de la hoja desde el icono de reloj.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab))

    labor_id = request.form.get('labor_id') or None
    lab = row_to_dict(execute('SELECT * FROM hoja_labores WHERE id=? AND hoja_id=?', (labor_id, hoja_id), fetchone=True)) if labor_id else None
    if str(h.get('estado') or '').upper() == 'ENVIADA':
        flash('No se puede registrar: la hoja ya fue enviada.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id or ''))
    if not lab:
        flash('Primero debes seleccionar una labor válida.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab='labores'))
    labor = (lab or h).get('labor')
    grupo = (lab or h).get('grupo')
    turno = limpiar_texto(request.form.get('turno') or (lab or h).get('turno') or 'DIA')
    tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or (lab or h).get('tipo_tareo') or 'JORNAL')
    try:
        horas = float(request.form.get('horas') or 0)
        cantidad = float(request.form.get('cantidad') or 0)
        a_noct = 0.0
    except Exception:
        flash('Horas / avance inválido.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab))
    hora_inicio = request.form.get('hora_inicio') or h.get('hora_inicio_default') or ('22:00' if turno == 'NOCHE' else '06:30')
    hora_fin = request.form.get('hora_fin') or h.get('hora_fin_default') or ('06:00' if turno == 'NOCHE' else '16:30')
    ref_inicio = request.form.get('ref_inicio') or h.get('ref_inicio_default') or '12:00'
    ref_fin = request.form.get('ref_fin') or h.get('ref_fin_default') or '13:00'
    horas_noct = 0
    if tab == 'trabajadores':
        coherente, msg_coh = horario_coherente(hora_inicio, hora_fin, ref_inicio, ref_fin)
        if not coherente:
            flash(msg_coh, 'danger')
            return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id or ''))
        horas = calcular_horas_laborales(hora_inicio, hora_fin, ref_inicio, ref_fin)
        horas_noct = calcular_horas_nocturnas(hora_inicio, hora_fin, ref_inicio, ref_fin)
    unidad = limpiar_texto(request.form.get('unidad') or ('BALDE' if tab == 'rendimiento' else tipo_tareo))
    metodo = limpiar_texto(request.form.get('metodo') or 'DIGITACIÓN')

    ok = 0; no_encontrados = []
    for dni in dnis:
        t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
        if not t:
            no_encontrados.append(dni)
            continue
        if tab == 'trabajadores' and scalar('SELECT COUNT(*) AS c FROM tareos WHERE hoja_id=? AND labor_id=? AND dni=?', (hoja_id, labor_id, dni)):
            no_encontrados.append(dni + ' duplicado')
            continue
        if tab == 'rendimiento' and not scalar('SELECT COUNT(*) AS c FROM tareos WHERE hoja_id=? AND labor_id=? AND dni=?', (hoja_id, labor_id, dni)):
            no_encontrados.append(dni + ' no registrado en Trabajadores de esta labor')
            continue
        h_reg = horas
        if tab == 'rendimiento':
            item = next((x for x in avance_pre_items if x.get('dni') == dni), None) if avance_pre_items else None
            cant_reg = float(item.get('cantidad') if item else cantidad)
            metodo_reg = item.get('metodo') if item else metodo
            execute('INSERT INTO lecturas_balde(hoja_id,labor_id,dni,trabajador,fecha_hora,a_diurno,a_noct,metodo,registrado_por) VALUES(?,?,?,?,?,?,?,?,?)',
                    (hoja_id,labor_id,dni,t.get('trabajador',''),now_str(),cant_reg,0.0,metodo_reg,session.get('usuario')), commit=True)
            ok += 1
            continue
        execute('''INSERT INTO tareos(hoja_id,labor_id,dni,trabajador,empresa,area,cargo,fecha,labor,lote,fundo,horas,cantidad,unidad,observacion,registrado_por,creado_en,hora_inicio,hora_fin,ref_inicio,ref_fin,turno,tipo_tareo,horas_nocturnas)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (hoja_id,labor_id,dni,t.get('trabajador',''),t.get('empresa',''),t.get('area',''),t.get('cargo',''),h.get('fecha'),labor,limpiar_texto(request.form.get('lote') or grupo),grupo,h_reg,cantidad,unidad,'',session.get('usuario'),now_str(),hora_inicio,hora_fin,ref_inicio,ref_fin,turno,tipo_tareo,horas_noct), commit=True)
        ok += 1
    msg = f'Registro guardado correctamente. Registros guardados: {ok}.'
    if no_encontrados:
        msg += ' No encontrados: ' + ', '.join(no_encontrados)
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for('detalle_hoja', hoja_id=hoja_id, tab=tab, labor_id=labor_id or ''))




@app.route('/tareo/<int:tareo_id>/editar-horas-form', methods=['GET'])
@login_required
def editar_horas_tareo_form(tareo_id):
    r = row_to_dict(execute('SELECT * FROM tareos WHERE id=?', (tareo_id,), fetchone=True))
    if not r:
        flash('Registro de trabajador no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (r.get('hoja_id'),), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))
    hi = r.get('hora_inicio') or ('22:00' if r.get('turno') == 'NOCHE' else '06:30')
    hf = r.get('hora_fin') or ('06:00' if r.get('turno') == 'NOCHE' else '16:30')
    ri = r.get('ref_inicio') or '12:00'
    rf = r.get('ref_fin') or '13:00'
    body = f'''
    <div class="phone-wrap">
      <div class="page-card p-3">
        <div class="d-flex align-items-center justify-content-between mb-2">
          <h5 class="fw-bold text-success m-0"><i class="bi bi-pencil-square"></i> Modificar horario</h5>
          <a class="back-mini" href="{url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or '')}">×</a>
        </div>
        <div class="alert alert-light border small mb-2">
          Trabajador: <b>{r.get('trabajador') or ''}</b><br>DNI: <b>{r.get('dni') or ''}</b>
        </div>
        <form method="post" action="{url_for('editar_horas_tareo', tareo_id=tareo_id)}" id="frmEditTareoStandalone">
          <div class="row g-2">
            <div class="col-6"><label class="form-label">Hora inicio</label><input name="hora_inicio" class="form-control" value="{hi}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
            <div class="col-6"><label class="form-label">Hora fin</label><input name="hora_fin" class="form-control" value="{hf}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
            <div class="col-6"><label class="form-label">Ref. inicio</label><input name="ref_inicio" class="form-control" value="{ri}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
            <div class="col-6"><label class="form-label">Ref. fin</label><input name="ref_fin" class="form-control" value="{rf}" required pattern="^([01]?[0-9]|2[0-3]):[0-5][0-9]$"></div>
          </div>
          <div class="field-help mt-2">Se permiten horarios 24 horas y cruce de medianoche. El refrigerio debe estar dentro de la jornada.</div>
          <button class="btn btn-green w-100 mt-3" type="submit">GUARDAR CAMBIOS</button>
        </form>
      </div>
    </div>'''
    return render_template_string(BASE_HTML, title='Modificar horario', body=body)

@app.route('/tareo/<int:tareo_id>/editar-horas', methods=['POST'])
@login_required
def editar_horas_tareo(tareo_id):
    r = row_to_dict(execute('SELECT * FROM tareos WHERE id=?', (tareo_id,), fetchone=True))
    if not r:
        flash('Registro de trabajador no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (r.get('hoja_id'),), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if str(h.get('estado') or '').upper() == 'ENVIADA':
        flash('No se puede editar: la hoja de tareo ya fue enviada.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))
    def okhora(v):
        return bool(re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', str(v or '').strip()))
    hi=(request.form.get('hora_inicio') or '').strip().zfill(5)
    hf=(request.form.get('hora_fin') or '').strip().zfill(5)
    ri=(request.form.get('ref_inicio') or '').strip().zfill(5)
    rf=(request.form.get('ref_fin') or '').strip().zfill(5)
    if not all(okhora(x) for x in [hi,hf,ri,rf]):
        flash('Horario inválido. Usa formato 24 horas HH:MM.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))
    coherente, msg_coh = horario_coherente(hi, hf, ri, rf)
    if not coherente:
        flash(msg_coh, 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))
    horas = calcular_horas_laborales(hi,hf,ri,rf)
    noct = calcular_horas_nocturnas(hi,hf,ri,rf)
    execute('UPDATE tareos SET hora_inicio=?, hora_fin=?, ref_inicio=?, ref_fin=?, horas=?, horas_nocturnas=? WHERE id=?', (hi,hf,ri,rf,horas,noct,tareo_id), commit=True)
    flash('Horas del trabajador actualizadas correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))


@app.route('/tareo/<int:tareo_id>/eliminar')
@login_required
def eliminar_tareo(tareo_id):
    r = row_to_dict(execute('SELECT * FROM tareos WHERE id=?', (tareo_id,), fetchone=True))
    if not r:
        flash('Registro no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if hoja_enviada(r.get('hoja_id')):
        flash('No se puede eliminar: la hoja ya fue enviada.', 'danger')
    else:
        execute('DELETE FROM tareos WHERE id=?', (tareo_id,), commit=True)
        flash('Trabajador eliminado del tareo.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=r.get('hoja_id'), tab='trabajadores', labor_id=r.get('labor_id') or ''))

@app.route('/lectura/<int:lectura_id>/eliminar')
@login_required
def eliminar_lectura(lectura_id):
    l = row_to_dict(execute('SELECT * FROM lecturas_balde WHERE id=?', (lectura_id,), fetchone=True))
    if not l:
        flash('Avance no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if hoja_enviada(l.get('hoja_id')):
        flash('No se puede eliminar: la hoja ya fue enviada.', 'danger')
    else:
        execute('DELETE FROM lecturas_balde WHERE id=?', (lectura_id,), commit=True)
        flash('Avance eliminado correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))

@app.route('/lectura/<int:lectura_id>/editar', methods=['POST'])
@login_required
def editar_lectura(lectura_id):
    l = row_to_dict(execute('SELECT * FROM lecturas_balde WHERE id=?', (lectura_id,), fetchone=True))
    if not l:
        flash('Avance no encontrado.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if hoja_enviada(l.get('hoja_id')):
        flash('No se puede editar: la hoja ya fue enviada.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))
    try:
        cant = float(request.form.get('cantidad') or 0)
        noct = float(request.form.get('a_noct') or 0)
    except Exception:
        flash('Cantidad inválida.', 'danger')
        return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))
    execute('UPDATE lecturas_balde SET a_diurno=?, a_noct=? WHERE id=?', (cant, noct, lectura_id), commit=True)
    flash('Avance actualizado correctamente.', 'success')
    return redirect(url_for('detalle_hoja', hoja_id=l.get('hoja_id'), tab='rendimiento', labor_id=l.get('labor_id') or ''))

@app.route('/hoja/<int:hoja_id>/editar', methods=['GET','POST'])
@login_required
def editar_hoja(hoja_id):
    h = row_to_dict(execute('SELECT * FROM hojas_tareo WHERE id=?', (hoja_id,), fetchone=True))
    if not h:
        flash('Hoja no encontrada.', 'danger')
        return redirect(url_for('hojas_tareo'))
    if request.method == 'POST':
        fecha = request.form.get('fecha') or h.get('fecha')
        grupo = limpiar_texto(request.form.get('actividad'))
        subgrupo = limpiar_texto(request.form.get('labor'))
        labor = limpiar_texto(request.form.get('consumidor'))
        responsable = limpiar_texto(request.form.get('responsable'))
        turno = limpiar_texto(request.form.get('turno') or 'DIA')
        tipo_tareo = limpiar_texto(request.form.get('tipo_tareo') or 'JORNAL')
        if not grupo or not subgrupo or not responsable:
            flash('Actividad, labor y responsable son obligatorios.', 'danger')
            return redirect(url_for('editar_hoja', hoja_id=hoja_id))
        execute('UPDATE hojas_tareo SET fecha=?,grupo=?,subgrupo=?,labor=?,responsable=?,turno=?,tipo_tareo=? WHERE id=?', (fecha,grupo,subgrupo,labor,responsable,turno,tipo_tareo,hoja_id), commit=True)
        execute('UPDATE hoja_labores SET grupo=?,subgrupo=?,labor=?,responsable=?,turno=?,tipo_tareo=? WHERE hoja_id=?', (grupo,subgrupo,labor,responsable,turno,tipo_tareo,hoja_id), commit=True)
        flash('Hoja modificada correctamente.', 'success')
        return redirect(url_for('hojas_tareo'))
    body = """
    <div class="phone-wrap desktop-pad"><h2 class="header-title">MODIFICAR HOJA</h2><div class="page-card"><form method="post" class="floating-card m-2">
      <a class="back-mini" href="{{url_for('hojas_tareo')}}"><i class="bi bi-chevron-left"></i></a>
      <label class="form-label mt-2">FECHA</label><input type="date" name="fecha" class="form-control mb-2" value="{{h.fecha}}" required>
      <label class="form-label">ACTIVIDAD</label><input name="actividad" class="form-control mb-2" value="{{h.grupo}}" required>
      <label class="form-label">LABOR</label><input name="labor" class="form-control mb-2" value="{{h.subgrupo}}" required>
      <label class="form-label">CONSUMIDOR</label><input name="consumidor" class="form-control mb-2" value="{{h.labor}}">
      <label class="form-label">RESPONSABLE</label><input name="responsable" class="form-control mb-2" value="{{h.responsable}}" required>
      <div class="row g-2 mb-3"><div class="col-6"><label class="form-label">TURNO</label><select name="turno" class="form-select"><option {{'selected' if h.turno=='DIA' else ''}}>DIA</option><option {{'selected' if h.turno=='NOCHE' else ''}}>NOCHE</option></select></div><div class="col-6"><label class="form-label">TIPO</label><select name="tipo_tareo" class="form-select"><option {{'selected' if h.tipo_tareo=='JORNAL' else ''}}>JORNAL</option><option {{'selected' if h.tipo_tareo=='RENDIMIENTO' else ''}}>RENDIMIENTO</option></select></div></div>
      <button class="btn btn-green w-100">GUARDAR CAMBIOS</button>
    </form></div></div>"""
    return render_page(body, h=h)

@app.route('/hoja/<int:hoja_id>/eliminar')
@login_required
def eliminar_hoja(hoja_id):
    execute('DELETE FROM lecturas_balde WHERE hoja_id=?', (hoja_id,), commit=True)
    execute('DELETE FROM tareos WHERE hoja_id=?', (hoja_id,), commit=True)
    execute('DELETE FROM hoja_labores WHERE hoja_id=?', (hoja_id,), commit=True)
    execute('DELETE FROM hojas_tareo WHERE id=?', (hoja_id,), commit=True)
    flash('Hoja eliminada correctamente.', 'success')
    return redirect(url_for('hojas_tareo'))

@app.route('/hoja/<int:hoja_id>/enviar')
@login_required
def enviar_hoja(hoja_id):
    execute("UPDATE hojas_tareo SET estado='ENVIADA' WHERE id=?", (hoja_id,), commit=True)
    flash('Hoja marcada como ENVIADA.', 'success')
    return redirect(url_for('hojas_tareo'))

# ========================= SOPORTE / CONFIG / SINC =========================
@app.route('/soporte')
@login_required
def soporte():
    return render_page('<div class="phone-wrap"><h2 class="header-title">SOPORTE</h2><div class="floating-card"><b>Canal de soporte</b><p class="small text-muted mb-2">Registra incidencias de sincronización, acceso o lectura de fotocheck.</p><textarea class="form-control" rows="4" placeholder="Describe el problema..."></textarea><a class="btn btn-green w-100 mt-3" href="{{url_for(\'home\')}}">ENVIAR / VOLVER</a></div></div>')

@app.route('/configuraciones')
@admin_required
def configuraciones():
    body = '<div class="phone-wrap"><a class="back-mini" href="{{url_for(\'home\')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">CONFIGURACIÓN</h2><div class="floating-card"><a class="btn btn-green w-100 mb-2" href="{{url_for(\'cargar_base\')}}"><i class="bi bi-people"></i> Cargar trabajadores</a><a class="btn btn-green w-100 mb-2" href="{{url_for(\'cargar_actividades\')}}"><i class="bi bi-diagram-3"></i> Actividades / Labores / Consumidores</a><a class="btn btn-outline-success w-100 mb-2" href="{{url_for(\'plantilla_trabajadores\')}}">Plantilla trabajadores</a><a class="btn btn-outline-success w-100 mb-2" href="{{url_for(\'plantilla_actividades\')}}">Plantilla actividades</a><a class="btn btn-outline-secondary w-100 mb-2" href="{{url_for(\'usuarios\')}}"><i class="bi bi-people"></i> Usuarios</a><a class="btn btn-outline-secondary w-100" href="{{url_for(\'home\')}}">Volver</a></div></div>'
    return render_page(body)

@app.route('/sincronizacion')
@login_required
def sincronizacion():
    total = scalar('SELECT COUNT(*) AS c FROM trabajadores')
    body = '<div class="phone-wrap"><h2 class="header-title">SINCRONIZACIÓN</h2><div class="panel-green"><i class="bi bi-arrow-repeat"></i><h4>TABLAS MAESTRAS</h4></div><div class="floating-card" style="margin:-20px 10px 0"><p><b>Trabajadores:</b> {{total}}</p><p><b>Última actualización:</b> {{now}}</p><a class="btn btn-green w-100" href="{{url_for(\'home\')}}">SINCRONIZADO</a></div></div>'
    return render_page(body, total=total, now=now_str())

# ========================= REPORTES / CARGA / USUARIOS =========================
@app.route('/reportes')
@login_required
def reportes():
    desde = request.args.get('desde') or today_str(); hasta = request.args.get('hasta') or today_str(); q=request.args.get('q','').strip()
    params=[desde,hasta]; where='WHERE fecha>=? AND fecha<=?'
    if q:
        like=f"%{q.upper()}%"; where += ' AND (dni LIKE ? OR UPPER(trabajador) LIKE ? OR UPPER(labor) LIKE ?)'; params += [like,like,like]
    tareos = rows_to_dict(execute(f'SELECT * FROM tareos {where} ORDER BY creado_en DESC LIMIT 500', params, fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad report-wrap"><div class="config-header"><a class="back-mini" href="{{url_for('home')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title mb-2">REPORTES TAREO</h2></div><form class="floating-card mb-2"><div class="row g-2"><div class="col-6"><input class="form-control" type="date" name="desde" value="{{desde}}"></div><div class="col-6"><input class="form-control" type="date" name="hasta" value="{{hasta}}"></div><div class="col-9"><input class="form-control" name="q" value="{{q}}" placeholder="DNI / trabajador / labor"></div><div class="col-3 d-grid"><button class="btn btn-green"><i class="bi bi-search"></i></button></div></div><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('exportar_tareos',desde=desde,hasta=hasta,q=q)}}">EXPORTAR EXCEL</a></form>{% for r in tareos %}<div class="worker-card"><div class="worker-title"><div>{{r.fecha}}<br><b>{{r.trabajador}}</b></div><div class="text-end">{{r.dni}}<br><b>{{r.labor}}</b></div></div><div class="worker-grid"><div><label>HORAS</label><div class="mini-input">{{r.horas}}</div></div><div><label>CANT.</label><div class="mini-input">{{r.cantidad}}</div></div><div><label>UNIDAD</label><div class="mini-input">{{r.unidad}}</div></div></div></div>{% else %}<div class="alert alert-light border text-center">Sin datos.</div>{% endfor %}</div>"""
    return render_page(body, desde=desde, hasta=hasta, q=q, tareos=tareos)

@app.route('/exportar/tareos')
@login_required
def exportar_tareos():
    desde=request.args.get('desde') or today_str(); hasta=request.args.get('hasta') or today_str(); q=request.args.get('q','').strip()
    params=[desde,hasta]; where='WHERE fecha>=? AND fecha<=?'
    if q:
        like=f"%{q.upper()}%"; where += ' AND (dni LIKE ? OR UPPER(trabajador) LIKE ? OR UPPER(labor) LIKE ?)'; params += [like,like,like]
    rows=rows_to_dict(execute(f'SELECT fecha,dni,trabajador,empresa,area,cargo,labor,fundo,lote,horas,cantidad,unidad,turno,tipo_tareo,hora_inicio,hora_fin,ref_inicio,ref_fin,observacion,registrado_por,creado_en FROM tareos {where} ORDER BY creado_en DESC', params, fetchall=True))
    headers=['FECHA','DNI','TRABAJADOR','EMPRESA','AREA','CARGO','LABOR','FUNDO','LOTE','HORAS','CANTIDAD','UNIDAD','TURNO','TIPO_TAREO','HORA_INICIO','HORA_FIN','REF_INICIO','REF_FIN','OBSERVACION','REGISTRADO_POR','CREADO_EN']
    return excel_response(headers, rows, f'tareos_{desde}_a_{hasta}.xlsx', 'TAREOS')

@app.route('/cargar-base', methods=['GET','POST'])
@admin_required
def cargar_base():
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith('.xlsx'):
            flash('Suba un Excel .xlsx válido.', 'danger'); return redirect(url_for('cargar_base'))
        wb = load_workbook(f, data_only=True, read_only=True); ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: flash('Excel vacío.', 'danger'); return redirect(url_for('cargar_base'))
        headers = [normalizar_columna(c) for c in rows[0]]
        if 'DNI' not in headers:
            flash('La plantilla debe tener la columna DNI.', 'danger'); return redirect(url_for('cargar_base'))
        colmap = {'TRABAJADOR':'trabajador','NOMBRE':'trabajador','APELLIDOS Y NOMBRES':'trabajador','EMPRESA':'empresa','AREA':'area','ÁREA':'area','CARGO':'cargo','ACTIVIDAD':'actividad','PLANILLA':'planilla','ESTADO':'estado'}
        ins=upd=omi=0; conn=get_conn(); cur=conn.cursor(); ahora=now_str()
        for row in rows[1:]:
            item={headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            dni=limpiar_dni(item.get('DNI'))
            if len(dni)!=8: omi+=1; continue
            data={'trabajador':'','empresa':'','area':'','cargo':'','actividad':'','planilla':'','estado':'ACTIVO'}
            for col,key in colmap.items():
                if col in headers and item.get(col) not in (None,''): data[key]=limpiar_texto(item.get(col))
            cur.execute(qmark('SELECT id FROM trabajadores WHERE dni=?'), (dni,))
            if cur.fetchone():
                cur.execute(qmark('UPDATE trabajadores SET trabajador=?,empresa=?,area=?,cargo=?,actividad=?,planilla=?,estado=?,fecha_carga=? WHERE dni=?'), (data['trabajador'],data['empresa'],data['area'],data['cargo'],data['actividad'],data['planilla'],data['estado'],ahora,dni)); upd+=1
            else:
                cur.execute(qmark('INSERT INTO trabajadores(dni,trabajador,empresa,area,cargo,actividad,planilla,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?)'), (dni,data['trabajador'],data['empresa'],data['area'],data['cargo'],data['actividad'],data['planilla'],data['estado'],ahora)); ins+=1
        conn.commit(); cur.close(); conn.close(); flash(f'Carga completa. Insertados: {ins} | Actualizados: {upd} | Omitidos: {omi}', 'success')
        return redirect(url_for('cargar_base'))
    trabajadores = rows_to_dict(execute('SELECT * FROM trabajadores ORDER BY fecha_carga DESC, trabajador LIMIT 100', fetchall=True))
    body = """
    <div class="phone-wrap desktop-pad"><div class="config-header"><a class="back-mini" href="{{url_for('configuraciones')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">CONFIGURACIÓN – TRABAJADORES</h2></div><form method="post" enctype="multipart/form-data" class="floating-card mb-2"><label class="form-label">Archivo Excel .xlsx</label><input class="form-control mb-2" type="file" name="archivo" accept=".xlsx" required><button class="btn btn-green w-100">CARGAR BASE</button><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('plantilla_trabajadores')}}">PLANTILLA EXCEL</a></form>{% for r in trabajadores %}<div class="worker-card"><div class="worker-title"><div>{{r.dni}}<br><b>{{r.trabajador}}</b></div><div class="text-end">{{r.empresa}}<br><b>{{r.cargo}}</b></div></div></div>{% endfor %}</div>"""
    return render_page(body, trabajadores=trabajadores)

@app.route('/plantilla-trabajadores')
@admin_required
def plantilla_trabajadores():
    headers=['DNI','TRABAJADOR','EMPRESA','CARGO','ESTADO','FECHA']
    rows=[{'DNI':'12345678','TRABAJADOR':'APELLIDOS Y NOMBRES','EMPRESA':'AQUANQA I','CARGO':'OPERARIO','ESTADO':'ACTIVO','FECHA':today_str()}]
    return excel_response(headers, rows, 'plantilla_trabajadores.xlsx', 'TRABAJADORES')


@app.route('/cargar-actividades', methods=['GET','POST'])
@admin_required
def cargar_actividades():
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith('.xlsx'):
            flash('Suba un Excel .xlsx válido.', 'danger'); return redirect(url_for('cargar_actividades'))
        wb = load_workbook(f, data_only=True, read_only=True); ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('Excel vacío.', 'danger'); return redirect(url_for('cargar_actividades'))
        headers = [normalizar_columna(c) for c in rows[0]]
        aliases = {
            'COD. ACTIVIDAD':'cod_actividad','COD ACTIVIDAD':'cod_actividad','COD_ACTIVIDAD':'cod_actividad',
            'DESCRIPCION ACTIVIDAD':'desc_actividad','DESCRIPCIÓN ACTIVIDAD':'desc_actividad','DESC ACTIVIDAD':'desc_actividad','DESC. ACTIVIDAD':'desc_actividad','ACTIVIDAD':'desc_actividad',
            'COD. LABOR':'cod_labor','COD LABOR':'cod_labor','COD_LABOR':'cod_labor',
            'DESCRIPCION LABOR':'desc_labor','DESCRIPCIÓN LABOR':'desc_labor','DESC LABOR':'desc_labor','DESC. LABOR':'desc_labor','LABOR':'desc_labor',
            'COD. CONSUMIDOR':'cod_consumidor','COD CONSUMIDOR':'cod_consumidor','COD_CONSUMIDOR':'cod_consumidor',
            'DESCRIPCION CONSUMIDOR':'desc_consumidor','DESCRIPCIÓN CONSUMIDOR':'desc_consumidor','DESC CONSUMIDOR':'desc_consumidor','DESC. CONSUMIDOR':'desc_consumidor','CONSUMIDOR':'desc_consumidor','ZONA':'desc_consumidor','CAMPO':'desc_consumidor'
        }
        required = ['desc_actividad','desc_labor']
        mapped = {i: aliases.get(h) for i,h in enumerate(headers)}
        if not any(v=='desc_actividad' for v in mapped.values()) or not any(v=='desc_labor' for v in mapped.values()):
            flash('La plantilla debe tener Descripción Actividad y Descripción Labor.', 'danger'); return redirect(url_for('cargar_actividades'))
        conn=get_conn(); cur=conn.cursor(); ins=0; omi=0; ahora=now_str()
        # Carga tipo reemplazo: evita que al subir el mismo Excel se dupliquen miles de filas.
        cur.execute(qmark('DELETE FROM actividades_maestras'))
        vistos_import=set()
        for row in rows[1:]:
            data={'cod_actividad':'','desc_actividad':'','cod_labor':'','desc_labor':'','cod_consumidor':'','desc_consumidor':''}
            for i,val in enumerate(row):
                k=mapped.get(i)
                if k: data[k]=limpiar_texto(val)
            if not data['desc_actividad'] or not data['desc_labor']:
                omi += 1; continue
            key=(data['desc_actividad'],data['desc_labor'],data['desc_consumidor'])
            if key in vistos_import:
                omi += 1; continue
            vistos_import.add(key)
            cur.execute(qmark('INSERT INTO actividades_maestras(cod_actividad,desc_actividad,cod_labor,desc_labor,cod_consumidor,desc_consumidor,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?)'),
                        (data['cod_actividad'],data['desc_actividad'],data['cod_labor'],data['desc_labor'],data['cod_consumidor'],data['desc_consumidor'],'ACTIVO',ahora)); ins += 1
        conn.commit(); cur.close(); conn.close()
        flash(f'Actividades cargadas. Insertados: {ins} | Omitidos: {omi}', 'success')
        return redirect(url_for('cargar_actividades'))
    datos = get_actividades_maestras(2000)
    total_reg = scalar('SELECT COUNT(*) AS total FROM actividades_maestras')
    body = """
    <div class="phone-wrap desktop-pad"><a class="back-mini" href="{{url_for('configuraciones')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">ACTIVIDADES / LABORES / CONSUMIDORES</h2>
    <form method="post" enctype="multipart/form-data" class="floating-card mb-2"><label class="form-label">Archivo Excel .xlsx</label><input class="form-control mb-2" type="file" name="archivo" accept=".xlsx" required><button class="btn btn-green w-100">CARGAR ACTIVIDADES</button><a class="btn btn-outline-success w-100 mt-2" href="{{url_for('plantilla_actividades')}}">PLANTILLA ACTIVIDADES</a></form>
    <div class="alert alert-success border">Base cargada: {{total_reg}} filas. Mostrando combinaciones únicas ACTIVIDAD/LABOR/CONSUMIDOR para evitar repetidos.</div>
    {% for r in datos %}<div class="worker-card"><div class="worker-title"><div>ACTIVIDAD<br><b>{{r.desc_actividad}}</b></div><div class="text-end">LABOR<br><b>{{r.desc_labor}}</b></div></div><div class="small-label mt-2">CONSUMIDOR</div><div class="small-value">{{r.desc_consumidor or 'NO OBLIGATORIO'}}</div></div>{% else %}<div class="alert alert-light border text-center">Sin actividades cargadas.</div>{% endfor %}</div>"""
    return render_page(body, datos=datos, total_reg=total_reg)

@app.route('/plantilla-actividades')
@admin_required
def plantilla_actividades():
    headers=['Cod. Actividad','Descripción Actividad','Cod. Labor','Descripción Labor','Cod. Consumidor','Descripción Consumidor']
    rows=[{'Cod. Actividad':'ACT001','Descripción Actividad':'COSECHA','Cod. Labor':'LAB001','Descripción Labor':'COSECHA MANUAL','Cod. Consumidor':'CON001','Descripción Consumidor':'CAMPO 01'},
          {'Cod. Actividad':'ACT001','Descripción Actividad':'COSECHA','Cod. Labor':'LAB002','Descripción Labor':'COSECHA SELECTIVA','Cod. Consumidor':'','Descripción Consumidor':''}]
    return excel_response(headers, rows, 'plantilla_actividades_labores_consumidores.xlsx', 'ACTIVIDADES')

@app.route('/api/actividades-maestras')
@login_required
def api_actividades_maestras():
    """Devuelve el árbol Actividad/Labor/Consumidor de forma tolerante.
    Primero usa actividades_maestras; si no hay data, arma opciones con lo ya registrado en hojas/labores
    y con actividad/cargo de trabajadores. Así el modal nunca queda vacío.
    """
    data = []
    try:
        data.extend(get_actividades_maestras())
    except Exception:
        data = []
    try:
        for r in rows_to_dict(execute("SELECT DISTINCT grupo AS desc_actividad, subgrupo AS desc_labor, labor AS desc_consumidor FROM hoja_labores WHERE COALESCE(grupo,'')<>'' OR COALESCE(subgrupo,'')<>'' OR COALESCE(labor,'')<>'' LIMIT 1000", fetchall=True)):
            data.append(r)
    except Exception:
        pass
    try:
        for r in rows_to_dict(execute("SELECT DISTINCT grupo AS desc_actividad, subgrupo AS desc_labor, labor AS desc_consumidor FROM hojas_tareo WHERE COALESCE(grupo,'')<>'' OR COALESCE(subgrupo,'')<>'' OR COALESCE(labor,'')<>'' LIMIT 1000", fetchall=True)):
            data.append(r)
    except Exception:
        pass
    try:
        for r in rows_to_dict(execute("SELECT DISTINCT actividad AS desc_actividad, cargo AS desc_labor, area AS desc_consumidor FROM trabajadores WHERE COALESCE(actividad,'')<>'' OR COALESCE(cargo,'')<>'' OR COALESCE(area,'')<>'' LIMIT 1000", fetchall=True)):
            data.append(r)
    except Exception:
        pass
    clean=[]; seen=set()
    for x in data:
        da=limpiar_texto(x.get('desc_actividad') or x.get('actividad') or x.get('grupo') or x.get('cod_actividad') or '')
        dl=limpiar_texto(x.get('desc_labor') or x.get('labor') or x.get('subgrupo') or x.get('cod_labor') or '')
        dc=limpiar_texto(x.get('desc_consumidor') or x.get('consumidor') or x.get('zona') or x.get('campo') or x.get('cod_consumidor') or '')
        key=(da,dl,dc)
        if (da or dl or dc) and key not in seen:
            seen.add(key); clean.append({'desc_actividad':da,'desc_labor':dl,'desc_consumidor':dc,'cod_actividad':x.get('cod_actividad',''),'cod_labor':x.get('cod_labor',''),'cod_consumidor':x.get('cod_consumidor','')})
    if not clean:
        clean = [
            {'desc_actividad':'ADMINISTRACION','desc_labor':'LABOR ADMINISTRATIVA','desc_consumidor':'OFICINA CENTRAL','cod_actividad':'ACT000','cod_labor':'LAB000','cod_consumidor':'CON000'},
            {'desc_actividad':'ADMISION','desc_labor':'CONTROL DOCUMENTARIO','desc_consumidor':'OFICINA 01','cod_actividad':'ACT003','cod_labor':'LAB004','cod_consumidor':'CON004'},
            {'desc_actividad':'COSECHA','desc_labor':'COSECHA MANUAL','desc_consumidor':'CAMPO 01','cod_actividad':'ACT001','cod_labor':'LAB001','cod_consumidor':'CON001'},
            {'desc_actividad':'COSECHA','desc_labor':'COSECHA SELECTIVA','desc_consumidor':'CAMPO 02','cod_actividad':'ACT001','cod_labor':'LAB002','cod_consumidor':'CON002'},
            {'desc_actividad':'PODA','desc_labor':'PODA SANITARIA','desc_consumidor':'LOTE 01','cod_actividad':'ACT002','cod_labor':'LAB003','cod_consumidor':'CON003'},
        ]
    return jsonify(ok=True, data=clean, total=len(clean))

@app.route('/usuarios', methods=['GET','POST'])
@admin_required
def usuarios():
    if request.method == 'POST':
        usuario=request.form.get('usuario','').strip(); nombres=limpiar_texto(request.form.get('nombres')); clave=request.form.get('clave',''); rol=request.form.get('rol','operador'); estado=request.form.get('estado','ACTIVO')
        if not usuario or not clave: flash('Usuario y clave son obligatorios.', 'danger')
        else:
            try:
                execute('INSERT INTO usuarios(usuario,password_hash,nombres,rol,estado,creado_en) VALUES(?,?,?,?,?,?)', (usuario,generate_password_hash(clave),nombres,rol,estado,now_str()), commit=True); flash('Usuario creado.', 'success')
            except Exception as e: flash(f'No se pudo crear usuario: {e}', 'danger')
        return redirect(url_for('usuarios'))
    users=rows_to_dict(execute('SELECT usuario,nombres,rol,estado,creado_en FROM usuarios ORDER BY usuario', fetchall=True))
    body='<div class="phone-wrap"><a class="back-mini" href="{{url_for(\'configuraciones\')}}"><i class="bi bi-chevron-left"></i></a><h2 class="header-title">USUARIOS</h2><form method="post" class="floating-card mb-2"><input class="form-control mb-2" name="usuario" placeholder="Usuario" required><input class="form-control mb-2" name="nombres" placeholder="Nombres"><input class="form-control mb-2" type="password" name="clave" placeholder="Clave" required><select class="form-select mb-2" name="rol"><option value="operador">operador</option><option value="admin">admin</option></select><select class="form-select mb-2" name="estado"><option>ACTIVO</option><option>INACTIVO</option></select><button class="btn btn-green w-100">Guardar</button></form>{% for u in users %}<div class="worker-card"><b>{{u.usuario}}</b> · {{u.rol}}<br><span class="small text-muted">{{u.nombres}} · {{u.estado}}</span></div>{% endfor %}</div>'
    return render_page(body, users=users)

# ========================= COMPAT API MARCACIÓN BÁSICA =========================

@app.route('/api/trabajador-labor/<int:hoja_id>/<int:labor_id>/<dni>')
@login_required
def api_trabajador_labor(hoja_id, labor_id, dni):
    dni = limpiar_dni(dni)
    if len(dni) != 8:
        return jsonify(ok=False, msg='DNI inválido.')
    r = row_to_dict(execute('SELECT dni, trabajador FROM tareos WHERE hoja_id=? AND labor_id=? AND dni=? ORDER BY id DESC LIMIT 1', (hoja_id, labor_id, dni), fetchone=True))
    if not r:
        return jsonify(ok=False, msg='Debe registrar primero al trabajador en el módulo Trabajadores de esta labor.')
    return jsonify(ok=True, trabajador=r)

@app.route('/api/trabajador/<dni>')
@login_required
def api_trabajador(dni):
    dni = limpiar_dni(dni)
    if len(dni) != 8: return jsonify(ok=False, msg='DNI inválido.')
    t = row_to_dict(execute('SELECT * FROM trabajadores WHERE dni=?', (dni,), fetchone=True))
    if not t: return jsonify(ok=False, msg='DNI no encontrado en la base de trabajadores.')
    return jsonify(ok=True, trabajador=t, sugerido='ENTRADA')

# ========================= PWA =========================
@app.route('/manifest.webmanifest')
@app.route('/manifest.json')
def manifest():
    return jsonify({"name":"Tareo Móvil PRIZE","short_name":"Tareo","start_url":"/","display":"standalone","background_color":"#ffffff","theme_color":"#2f773b","icons":[]})

@app.route('/sw.js')
def sw():
    return Response("self.addEventListener('install',e=>self.skipWaiting()); self.addEventListener('fetch',e=>{});", mimetype='application/javascript')

try:
    init_db()
except Exception as e:
    print('ERROR inicializando base de datos:', e)

if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    app.run(host='0.0.0.0', port=port, debug=False)
